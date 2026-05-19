from flask import Flask, render_template, request, redirect, url_for, session, flash, g, jsonify, Response, send_file, send_from_directory
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from flask_wtf.csrf import CSRFProtect, CSRFError
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from authlib.integrations.flask_client import OAuth as _FlaskOAuth
from dotenv import load_dotenv
import logging
import base64
import io
import sqlite3
import os
import re
import hashlib
import secrets
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta
from functools import wraps
from urllib.parse import urlencode
try:
    import openpyxl
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

# ── Environment bootstrap ─────────────────────────────────────────────────────
# Three environments: development | staging | production
#
#   development  – local machine, SQLite, debug on
#   staging      – DigitalOcean staging app, PostgreSQL staging DB, debug off
#   production   – DigitalOcean production app, PostgreSQL prod DB, debug off
#
# Set APP_ENV in the platform dashboard (DigitalOcean) or in the appropriate
# .env.* file. Platform variables always win (override=False below).
_app_env = os.environ.get('APP_ENV', '')

if _app_env == 'production':
    load_dotenv('.env.production', override=False)
elif _app_env == 'staging':
    load_dotenv('.env.staging', override=False)
else:
    load_dotenv('.env.development')

# Re-read after load_dotenv so APP_ENV inside the file takes effect.
_app_env       = os.environ.get('APP_ENV', '')
_is_production = (_app_env == 'production')
_is_staging    = (_app_env == 'staging')

app = Flask(__name__)
logging.basicConfig(level=logging.INFO)
app.logger.setLevel(logging.INFO)

_secret_key = os.environ.get('SECRET_KEY', '')
if not _secret_key:
    import secrets as _secrets
    _secret_key = _secrets.token_hex(32)
    print('[WARNING] SECRET_KEY not set in .env — using a random key. '
          'Sessions will not persist across restarts.')
app.secret_key = _secret_key

# DEBUG=True shows detailed error pages; always False in production
app.debug = os.environ.get('DEBUG', 'False').lower() == 'true'

app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=_is_production,  # HTTPS-only cookies in production
    WTF_CSRF_TIME_LIMIT=3600,
)

csrf = CSRFProtect(app)

limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    storage_uri='memory://',
)

# ── OAuth (Google + LinkedIn) ─────────────────────────────────────────────────
_oauth = _FlaskOAuth(app)

_GOOGLE_CLIENT_ID     = os.environ.get('GOOGLE_CLIENT_ID', '')
_GOOGLE_CLIENT_SECRET = os.environ.get('GOOGLE_CLIENT_SECRET', '')
_LINKEDIN_CLIENT_ID     = os.environ.get('LINKEDIN_CLIENT_ID', '')
_LINKEDIN_CLIENT_SECRET = os.environ.get('LINKEDIN_CLIENT_SECRET', '')

if _GOOGLE_CLIENT_ID and _GOOGLE_CLIENT_SECRET:
    _oauth.register(
        'google',
        client_id=_GOOGLE_CLIENT_ID,
        client_secret=_GOOGLE_CLIENT_SECRET,
        server_metadata_url='https://accounts.google.com/.well-known/openid-configuration',
        client_kwargs={'scope': 'openid email profile'},
    )

# LinkedIn is handled with raw HTTP calls (no Authlib OIDC) to avoid
# server_metadata_url fetch failures and ID-token validation issues.

DATABASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'skillbasehire.db')

# ── PostgreSQL adapter ────────────────────────────────────────────────────────
# Use PostgreSQL when DATABASE_URL is available and not explicitly in dev mode.
# DigitalOcean injects DATABASE_URL automatically; set APP_ENV=development locally
# (in your shell or .env.development) to force SQLite even if DATABASE_URL is set.
_DATABASE_URL = os.environ.get('DATABASE_URL', '')
if _DATABASE_URL.startswith('postgres://'):
    _DATABASE_URL = _DATABASE_URL.replace('postgres://', 'postgresql://', 1)
_USE_POSTGRES = bool(_DATABASE_URL) and (_app_env != 'development')
print(f'[STARTUP] APP_ENV={_app_env!r}  _is_production={_is_production}  _is_staging={_is_staging}  '
      f'_USE_POSTGRES={_USE_POSTGRES}  DB={"postgres" if _USE_POSTGRES else "sqlite"}')

if _USE_POSTGRES:
    import psycopg2
    import psycopg2.extras
    import psycopg2.pool as _pgpool
    _pg_pool = None  # created lazily on first request

    def _get_pg_pool():
        global _pg_pool
        if _pg_pool is None:
            _pg_pool = _pgpool.SimpleConnectionPool(1, 10, _DATABASE_URL)
        return _pg_pool

    def _adapt_sql(sql):
        """Translate SQLite-flavoured SQL to PostgreSQL."""
        # Positional placeholder
        sql = sql.replace('?', '%s')
        # INSERT OR IGNORE → INSERT … ON CONFLICT DO NOTHING
        if re.search(r'INSERT\s+OR\s+IGNORE', sql, re.IGNORECASE):
            sql = re.sub(r'INSERT\s+OR\s+IGNORE\s+', 'INSERT ', sql, flags=re.IGNORECASE)
            sql = sql.rstrip().rstrip(';') + ' ON CONFLICT DO NOTHING'
        # GROUP_CONCAT(a.col||':'||b.col, sep) — mixed-type concat with integer column
        sql = re.sub(
            r"GROUP_CONCAT\(\s*(\w+\.\w+)\s*\|\|\s*':'\s*\|\|\s*(\w+\.\w+)\s*,\s*(',?')\)",
            lambda m: (
                f"STRING_AGG({m.group(1)}||':'||{m.group(2)}::TEXT, {m.group(3)})"
            ),
            sql, flags=re.IGNORECASE,
        )
        # GROUP_CONCAT(expr, sep) — simple column or expression
        def _gc(m):
            content = m.group(1)
            sep_m = re.search(r",\s*('(?:[^']|'')*')\s*$", content)
            if sep_m:
                expr = content[:sep_m.start()].strip()
                sep  = sep_m.group(1)
            else:
                expr, sep = content.strip(), "','"
            return f"STRING_AGG(CAST({expr} AS TEXT), {sep})"
        sql = re.sub(r'GROUP_CONCAT\(([^)]+)\)', _gc, sql, flags=re.IGNORECASE)
        return sql

    class _PGRow(dict):
        """psycopg2 RealDictRow wrapper that supports integer positional access."""
        def __getitem__(self, key):
            if isinstance(key, int):
                return list(self.values())[key]
            return super().__getitem__(key)

    class _PGCursor:
        def __init__(self, cur, has_returning=False):
            self._cur = cur
            self._has_returning = has_returning
            self._lastrowid = None
            self._rowid_fetched = False

        @property
        def lastrowid(self):
            if not self._rowid_fetched:
                self._rowid_fetched = True
                if self._has_returning:
                    row = self._cur.fetchone()
                    self._lastrowid = row['id'] if row else None
            return self._lastrowid

        def fetchone(self):
            row = self._cur.fetchone()
            return _PGRow(row) if row else None

        def fetchall(self):
            return [_PGRow(r) for r in (self._cur.fetchall() or [])]

        def __iter__(self):
            for row in self._cur:
                yield _PGRow(row)

    class _PGConn:
        def __init__(self, conn):
            self._conn = conn

        def execute(self, sql, params=None):
            pg_sql = _adapt_sql(sql)
            has_returning = False
            up = pg_sql.strip().upper()
            # Add RETURNING id to bare INSERTs so .lastrowid works
            if (up.startswith('INSERT') and
                    'ON CONFLICT' not in up and
                    'RETURNING' not in up):
                pg_sql = pg_sql.rstrip().rstrip(';') + ' RETURNING id'
                has_returning = True
            cur = self._conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cur.execute(pg_sql, params or ())
            return _PGCursor(cur, has_returning=has_returning)

        def executemany(self, sql, seq):
            cur = self._conn.cursor()
            cur.executemany(_adapt_sql(sql), seq)

        def executescript(self, sql):
            cur = self._conn.cursor()
            for stmt in [s.strip() for s in sql.split(';') if s.strip()]:
                try:
                    cur.execute(stmt)
                except Exception:
                    self._conn.rollback()
                    return

        def commit(self):
            self._conn.commit()

        def rollback(self):
            self._conn.rollback()

        def close(self):
            _get_pg_pool().putconn(self._conn)

    def _init_postgres_schema():
        import traceback as _tb
        try:
            import init_db as _schema
            conn = _get_pg_pool().getconn()
            try:
                cur = conn.cursor()
                # Tables: one transaction, all use IF NOT EXISTS
                for stmt in _schema.TABLE_STMTS:
                    cur.execute(stmt)
                cur.executemany(
                    "INSERT INTO skills (name, category, description) VALUES (%s, %s, %s)"
                    " ON CONFLICT (name) DO NOTHING",
                    _schema.SKILLS_DATA,
                )
                conn.commit()
                # Indexes: each in its own savepoint — guards against race condition
                # when multiple gunicorn workers run _init_postgres_schema concurrently
                for stmt in _schema.INDEX_STMTS:
                    try:
                        cur.execute('SAVEPOINT _idx')
                        cur.execute(stmt)
                        conn.commit()
                    except Exception:
                        cur.execute('ROLLBACK TO SAVEPOINT _idx')
                        conn.commit()
                print('[STARTUP] PostgreSQL tables/indexes ready')
            except Exception as e:
                conn.rollback()
                print(f'[STARTUP] Table init error: {e}')
                _tb.print_exc()

            # Migrations: each in its own savepoint so one failure doesn't block others
            applied = 0
            for stmt in _schema.MIGRATION_STMTS:
                try:
                    cur.execute('SAVEPOINT _mig')
                    cur.execute(stmt)
                    conn.commit()
                    applied += 1
                except Exception as e:
                    cur.execute('ROLLBACK TO SAVEPOINT _mig')
                    conn.commit()
                    print(f'[STARTUP] Migration skipped ({e}): {stmt[:60]}')
            print(f'[STARTUP] {applied}/{len(_schema.MIGRATION_STMTS)} migrations applied')
            _get_pg_pool().putconn(conn)
        except Exception as e:
            print(f'[STARTUP] Could not connect to PostgreSQL: {e}')
            _tb.print_exc()

    _init_postgres_schema()

PERSONAL_EMAIL_DOMAINS = {
    'gmail.com', 'yahoo.com', 'outlook.com', 'hotmail.com',
    'rediffmail.com', 'protonmail.com', 'icloud.com',
    'aol.com', 'ymail.com', 'mail.com', 'live.com',
}
RESUME_UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'uploads', 'resumes')
PHOTO_UPLOAD_FOLDER  = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'uploads', 'photos')
ALLOWED_RESUME_EXT = {'pdf', 'doc', 'docx'}
ALLOWED_PHOTO_EXT  = {'jpg', 'jpeg', 'png', 'webp'}
MAX_RESUME_SIZE = 1 * 1024 * 1024   # 1 MB
MAX_PHOTO_SIZE  = 100 * 1024         # 100 KB


class _SQLiteRow(dict):
    """Dict-based SQLite row that also supports integer positional access like sqlite3.Row."""
    def __getitem__(self, key):
        if isinstance(key, int):
            return list(self.values())[key]
        return super().__getitem__(key)


def _sqlite_row_factory(cur, row):
    return _SQLiteRow(zip([c[0] for c in cur.description], row))


def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        if _USE_POSTGRES:
            conn = _get_pg_pool().getconn()
            conn.autocommit = False
            db = g._database = _PGConn(conn)
        else:
            db = g._database = sqlite3.connect(DATABASE)
            db.row_factory = _sqlite_row_factory
    return db


@app.teardown_appcontext
def close_connection(_exception):
    db = getattr(g, '_database', None)
    if db is not None:
        if _exception and _USE_POSTGRES:
            try:
                db.rollback()
            except Exception:
                pass
        db.close()


@app.after_request
def set_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    # camera=(self) and microphone=(self) — allow same-origin pages (exam proctoring)
    # to request camera/mic access while blocking any embedded third-party iframes.
    response.headers['Permissions-Policy'] = 'geolocation=(), camera=(self), microphone=(self)'
    if _is_production:  # HSTS only makes sense over HTTPS in production
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    return response


@app.errorhandler(CSRFError)
def csrf_error(e):
    if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'error': 'Session expired. Please refresh and try again.'}), 400
    flash('Your session has expired. Please try again.', 'error')
    return redirect(request.referrer or url_for('home'))


@app.errorhandler(429)
def ratelimit_error(e):
    if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'error': 'Too many requests. Please wait and try again.'}), 429
    flash('Too many attempts. Please wait a moment and try again.', 'error')
    return redirect(request.referrer or url_for('home')), 429


def validate_file_magic(stream, allowed_types):
    """Return True if the file's magic bytes match one of the allowed_types."""
    header = stream.read(12)
    stream.seek(0)
    if 'pdf' in allowed_types and header[:4] == b'%PDF':
        return True
    if 'docx' in allowed_types and header[:2] == b'PK':
        return True
    if 'doc' in allowed_types and header[:4] == b'\xd0\xcf\x11\xe0':
        return True
    if ('jpg' in allowed_types or 'jpeg' in allowed_types) and header[:3] == b'\xff\xd8\xff':
        return True
    if 'png' in allowed_types and header[:4] == b'\x89PNG':
        return True
    if 'webp' in allowed_types and header[:4] == b'RIFF' and header[8:12] == b'WEBP':
        return True
    return False


def init_db():
    with app.app_context():
        db = get_db()
        db.executescript('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                email TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL CHECK(role IN ('candidate','recruiter')),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS candidate_profiles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER UNIQUE NOT NULL,
                headline TEXT DEFAULT '',
                location TEXT DEFAULT '',
                bio TEXT DEFAULT '',
                linkedin TEXT DEFAULT '',
                github TEXT DEFAULT '',
                FOREIGN KEY (user_id) REFERENCES users(id)
            );
            CREATE TABLE IF NOT EXISTS recruiter_profiles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER UNIQUE NOT NULL,
                company TEXT NOT NULL,
                company_bio TEXT DEFAULT '',
                website TEXT DEFAULT '',
                FOREIGN KEY (user_id) REFERENCES users(id)
            );
            CREATE TABLE IF NOT EXISTS skills (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                category TEXT NOT NULL,
                description TEXT DEFAULT ''
            );
            CREATE TABLE IF NOT EXISTS user_skills (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                skill_id INTEGER NOT NULL,
                verified INTEGER DEFAULT 0,
                score INTEGER DEFAULT 0,
                added_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(user_id, skill_id),
                FOREIGN KEY (user_id) REFERENCES users(id),
                FOREIGN KEY (skill_id) REFERENCES skills(id)
            );
            CREATE TABLE IF NOT EXISTS jobs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                recruiter_id INTEGER NOT NULL,
                title TEXT NOT NULL,
                company TEXT NOT NULL,
                location TEXT NOT NULL,
                job_type TEXT NOT NULL DEFAULT 'Full-time',
                description TEXT NOT NULL,
                requirements TEXT DEFAULT '',
                salary_min INTEGER DEFAULT 0,
                salary_max INTEGER DEFAULT 0,
                active INTEGER DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (recruiter_id) REFERENCES users(id)
            );
            CREATE TABLE IF NOT EXISTS job_skills (
                job_id INTEGER NOT NULL,
                skill_id INTEGER NOT NULL,
                PRIMARY KEY (job_id, skill_id),
                FOREIGN KEY (job_id) REFERENCES jobs(id),
                FOREIGN KEY (skill_id) REFERENCES skills(id)
            );
            CREATE TABLE IF NOT EXISTS applications (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                job_id INTEGER NOT NULL,
                candidate_id INTEGER NOT NULL,
                status TEXT DEFAULT 'applied',
                cover_letter TEXT DEFAULT '',
                applied_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(job_id, candidate_id),
                FOREIGN KEY (job_id) REFERENCES jobs(id),
                FOREIGN KEY (candidate_id) REFERENCES users(id)
            );
            CREATE TABLE IF NOT EXISTS candidate_work_experience (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                company TEXT NOT NULL,
                designation TEXT NOT NULL,
                start_date TEXT DEFAULT '',
                end_date TEXT DEFAULT '',
                is_current INTEGER DEFAULT 0,
                description TEXT DEFAULT '',
                FOREIGN KEY (user_id) REFERENCES users(id)
            );
            CREATE TABLE IF NOT EXISTS candidate_education (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                degree TEXT NOT NULL,
                college TEXT NOT NULL,
                start_year TEXT DEFAULT '',
                end_year TEXT DEFAULT '',
                FOREIGN KEY (user_id) REFERENCES users(id)
            );
            CREATE TABLE IF NOT EXISTS candidate_certifications (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                cert_name TEXT NOT NULL,
                issued_by TEXT DEFAULT '',
                year TEXT DEFAULT '',
                credential_url TEXT DEFAULT '',
                FOREIGN KEY (user_id) REFERENCES users(id)
            );
            CREATE TABLE IF NOT EXISTS candidate_projects (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                project_name TEXT NOT NULL,
                domain TEXT DEFAULT '',
                description TEXT DEFAULT '',
                project_url TEXT DEFAULT '',
                FOREIGN KEY (user_id) REFERENCES users(id)
            );
            CREATE TABLE IF NOT EXISTS notifications (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                candidate_id INTEGER NOT NULL,
                job_id INTEGER,
                application_id INTEGER,
                title TEXT NOT NULL,
                message TEXT NOT NULL,
                type TEXT DEFAULT 'general',
                is_read INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (candidate_id) REFERENCES users(id),
                FOREIGN KEY (job_id) REFERENCES jobs(id),
                FOREIGN KEY (application_id) REFERENCES applications(id)
            );
            CREATE TABLE IF NOT EXISTS skill_questions (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                skill_id       INTEGER NOT NULL,
                question_text  TEXT NOT NULL,
                option_a       TEXT NOT NULL,
                option_b       TEXT NOT NULL,
                option_c       TEXT NOT NULL,
                option_d       TEXT NOT NULL,
                correct_option TEXT NOT NULL CHECK(correct_option IN ('A','B','C','D')),
                marks          INTEGER DEFAULT 1,
                difficulty     TEXT DEFAULT 'Medium' CHECK(difficulty IN ('Easy','Medium','Hard')),
                created_by     INTEGER,
                created_at     TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (skill_id) REFERENCES skills(id),
                FOREIGN KEY (created_by) REFERENCES users(id)
            );
        ''')

        skills_data = [
            ('Python', 'Programming', 'General-purpose programming language'),
            ('JavaScript', 'Programming', 'Web scripting language'),
            ('Java', 'Programming', 'Object-oriented programming language'),
            ('TypeScript', 'Programming', 'Typed superset of JavaScript'),
            ('C++', 'Programming', 'Systems programming language'),
            ('Go', 'Programming', 'Compiled systems language'),
            ('Rust', 'Programming', 'Memory-safe systems language'),
            ('React', 'Frontend', 'JavaScript UI library'),
            ('Vue.js', 'Frontend', 'Progressive JavaScript framework'),
            ('Angular', 'Frontend', 'TypeScript-based web framework'),
            ('HTML/CSS', 'Frontend', 'Web markup and styling'),
            ('Tailwind CSS', 'Frontend', 'Utility-first CSS framework'),
            ('Node.js', 'Backend', 'JavaScript runtime'),
            ('Django', 'Backend', 'Python web framework'),
            ('Flask', 'Backend', 'Lightweight Python web framework'),
            ('FastAPI', 'Backend', 'Modern Python API framework'),
            ('Spring Boot', 'Backend', 'Java web framework'),
            ('SQL', 'Database', 'Structured query language'),
            ('PostgreSQL', 'Database', 'Advanced open-source database'),
            ('MySQL', 'Database', 'Popular relational database'),
            ('MongoDB', 'Database', 'Document-oriented NoSQL database'),
            ('Redis', 'Database', 'In-memory data structure store'),
            ('AWS', 'Cloud', 'Amazon Web Services'),
            ('Azure', 'Cloud', 'Microsoft Azure'),
            ('GCP', 'Cloud', 'Google Cloud Platform'),
            ('Docker', 'DevOps', 'Containerization platform'),
            ('Kubernetes', 'DevOps', 'Container orchestration'),
            ('CI/CD', 'DevOps', 'Continuous integration and deployment'),
            ('Machine Learning', 'AI/ML', 'Statistical learning algorithms'),
            ('Deep Learning', 'AI/ML', 'Neural network-based learning'),
            ('Data Analysis', 'Data', 'Statistical data examination'),
            ('Data Visualization', 'Data', 'Graphical data representation'),
            ('Selenium', 'Testing', 'Web automation testing'),
            ('Jest', 'Testing', 'JavaScript testing framework'),
            ('PyTest', 'Testing', 'Python testing framework'),
            ('REST APIs', 'Backend', 'RESTful API development'),
            ('GraphQL', 'Backend', 'Query language for APIs'),
            ('Git', 'Tools', 'Version control system'),
            ('Linux', 'Tools', 'Open-source operating system'),
            ('Agile/Scrum', 'Methodology', 'Agile project management'),
            ('Playwright', 'Testing', 'Modern end-to-end browser testing framework'),
            ('Appium', 'Testing', 'Mobile app automation testing framework'),
            ('API Testing', 'Testing', 'REST and GraphQL API testing'),
            ('Manual Testing', 'Testing', 'Manual software quality assurance'),
            ('Cypress', 'Testing', 'Fast, reliable end-to-end testing for the web'),
            ('TestNG', 'Testing', 'Java testing framework inspired by JUnit'),
        ]
        db.executemany(
            'INSERT OR IGNORE INTO skills (name, category, description) VALUES (?, ?, ?)',
            skills_data
        )

        # Schema migrations — safe, idempotent
        for stmt in [
            "ALTER TABLE users ADD COLUMN email_verified INTEGER DEFAULT 0",
            "ALTER TABLE users ADD COLUMN email_verified_at TIMESTAMP",
            "ALTER TABLE users ADD COLUMN email_verification_token_hash TEXT",
            "ALTER TABLE users ADD COLUMN email_verification_expires_at TIMESTAMP",
            "ALTER TABLE users ADD COLUMN email_verification_sent_at TIMESTAMP",
            "ALTER TABLE candidate_profiles ADD COLUMN phone TEXT DEFAULT ''",
            "ALTER TABLE candidate_profiles ADD COLUMN job_title TEXT DEFAULT ''",
            "ALTER TABLE candidate_profiles ADD COLUMN experience TEXT DEFAULT ''",
            "ALTER TABLE candidate_profiles ADD COLUMN resume_filename TEXT",
            "ALTER TABLE candidate_profiles ADD COLUMN work_status TEXT DEFAULT ''",
            "ALTER TABLE candidate_profiles ADD COLUMN work_mode TEXT DEFAULT ''",
            "ALTER TABLE candidate_profiles ADD COLUMN notice_period TEXT DEFAULT ''",
            "ALTER TABLE candidate_profiles ADD COLUMN expected_salary TEXT DEFAULT ''",
            "ALTER TABLE candidate_profiles ADD COLUMN willing_to_relocate INTEGER DEFAULT 0",
            "ALTER TABLE recruiter_profiles ADD COLUMN phone TEXT DEFAULT ''",
            "ALTER TABLE recruiter_profiles ADD COLUMN job_title TEXT DEFAULT ''",
            "ALTER TABLE recruiter_profiles ADD COLUMN company_size TEXT DEFAULT ''",
            "ALTER TABLE recruiter_profiles ADD COLUMN industry TEXT DEFAULT ''",
            "ALTER TABLE recruiter_profiles ADD COLUMN company_location TEXT DEFAULT ''",
            "ALTER TABLE candidate_projects ADD COLUMN year TEXT DEFAULT ''",
            "ALTER TABLE jobs ADD COLUMN skill_test_mandatory INTEGER DEFAULT 0",
            "ALTER TABLE candidate_profiles ADD COLUMN gender TEXT DEFAULT ''",
            "ALTER TABLE candidate_profiles ADD COLUMN differently_abled TEXT DEFAULT 'no'",
            "ALTER TABLE candidate_profiles ADD COLUMN profile_photo TEXT",
            "ALTER TABLE recruiter_profiles ADD COLUMN profile_photo TEXT",
            "ALTER TABLE applications ADD COLUMN updated_at TIMESTAMP",
            "ALTER TABLE notifications ADD COLUMN redirect_url TEXT DEFAULT ''",
            "ALTER TABLE user_skills ADD COLUMN correct_answers INTEGER",
            "ALTER TABLE user_skills ADD COLUMN time_taken_secs INTEGER",
            "ALTER TABLE users ADD COLUMN google_id TEXT DEFAULT NULL",
            "ALTER TABLE users ADD COLUMN linkedin_id TEXT DEFAULT NULL",
            "ALTER TABLE skill_questions ADD COLUMN experience_level TEXT DEFAULT 'All'",
            "ALTER TABLE users ADD COLUMN password_reset_token_hash TEXT DEFAULT NULL",
            "ALTER TABLE users ADD COLUMN password_reset_expires_at TIMESTAMP DEFAULT NULL",
        ]:
            try:
                db.execute(stmt)
            except Exception:
                pass
        db.commit()


# ── Context processor ───────────────────────────────────────────────────────

@app.context_processor
def inject_notif_count():
    if session.get('role') == 'candidate' and session.get('user_id'):
        try:
            db = get_db()
            count = db.execute(
                'SELECT COUNT(*) FROM notifications WHERE candidate_id=? AND is_read=0',
                [session['user_id']]
            ).fetchone()[0]
            return {'unread_notif_count': count}
        except Exception:
            pass
    return {'unread_notif_count': 0}


# ── Jinja2 filters ──────────────────────────────────────────────────────────

@app.template_filter('timeago')
def timeago_filter(date_str):
    if not date_str:
        return ''
    try:
        if isinstance(date_str, str):
            dt = datetime.strptime(str(date_str)[:19], '%Y-%m-%d %H:%M:%S')
        else:
            dt = date_str
        diff = datetime.now() - dt
        if diff.days == 0:
            hours = diff.seconds // 3600
            return 'Just now' if hours == 0 else f'{hours}h ago'
        elif diff.days == 1:
            return 'Yesterday'
        elif diff.days < 7:
            return f'{diff.days}d ago'
        elif diff.days < 30:
            return f'{diff.days // 7}w ago'
        else:
            return dt.strftime('%b %d, %Y')
    except Exception:
        return str(date_str)


@app.template_filter('salary')
def salary_filter(job):
    mn, mx = job['salary_min'], job['salary_max']
    if not mn and not mx:
        return 'Negotiable'
    if mn and mx:
        return f'₹{mn // 1000}K – ₹{mx // 1000}K'
    if mn:
        return f'₹{mn // 1000}K+'
    return f'Up to ₹{mx // 1000}K'


# ── Auth helpers ─────────────────────────────────────────────────────────────

def _auth_role_check(required_role):
    """Verify session AND database agree on the user's role.
    Returns (allowed: bool, db_role: str|None).
    Caches the DB lookup on g for the duration of the request.
    """
    area  = request.endpoint or request.path
    uid   = session.get('user_id')
    email = session.get('email', 'unknown')

    if not uid:
        app.logger.warning('[AUTH_ROLE_CHECK] email=%s stored_role=none requested_area=%s allowed=false',
                           email, area)
        return False, None

    if not hasattr(g, '_auth_db_user'):
        try:
            g._auth_db_user = get_db().execute(
                'SELECT id, role, email FROM users WHERE id=?', [uid]
            ).fetchone()
        except Exception as exc:
            app.logger.error('[AUTH_ROLE_CHECK] DB lookup failed uid=%s: %s', uid, exc)
            g._auth_db_user = None

    db_user  = g._auth_db_user
    db_role  = db_user['role'] if db_user else 'none'
    db_email = (db_user['email'] if db_user else None) or email
    allowed  = db_user is not None and db_role == required_role

    app.logger.info('[AUTH_ROLE_CHECK] email=%s stored_role=%s requested_area=%s allowed=%s',
                    db_email, db_role, area, 'true' if allowed else 'false')
    return allowed, db_role


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to continue.', 'error')
            return redirect(url_for('home'))
        return f(*args, **kwargs)
    return decorated


def candidate_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('candidate_login'))
        allowed, db_role = _auth_role_check('candidate')
        if not allowed:
            session.clear()
            flash('Access denied. This area is for candidates only.', 'error')
            return redirect(url_for('candidate_login'))
        return f(*args, **kwargs)
    return decorated


def recruiter_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('recruiter_login'))
        allowed, db_role = _auth_role_check('recruiter')
        if not allowed:
            session.clear()
            flash('Access denied. This area is for recruiters only.', 'error')
            return redirect(url_for('recruiter_login'))
        return f(*args, **kwargs)
    return decorated


_SUPER_ADMIN_EMAIL = 'admin@skillbasehire.com'
_ADMIN_TOKEN       = os.environ.get('ADMIN_MANAGE_QUESTIONS_TOKEN', '')

# ── Email sender identity ─────────────────────────────────────────────────────
_SMTP_FROM_EMAIL = os.environ.get('SMTP_FROM_EMAIL', 'noreply@skillbasehire.com')
_SMTP_FROM_NAME  = os.environ.get('SMTP_FROM_NAME',  'SkillBaseHire')


def admin_required(f):
    """Restricts access to the super-admin recruiter account only."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('recruiter_login'))
        allowed, db_role = _auth_role_check('recruiter')
        if not allowed:
            session.clear()
            app.logger.warning('[AUTH_ROLE_CHECK] admin_required blocked email=%s db_role=%s',
                               session.get('email'), db_role)
            flash('Access denied.', 'error')
            return redirect(url_for('recruiter_login'))
        if session.get('email', '').lower() != _SUPER_ADMIN_EMAIL:
            area = request.endpoint or request.path
            app.logger.warning('[AUTH_ROLE_CHECK] admin_required non-admin email=%s requested_area=%s allowed=false',
                               session.get('email'), area)
            flash('You do not have permission to access this page.', 'error')
            return redirect(url_for('recruiter_dashboard'))
        return f(*args, **kwargs)
    return decorated


_EMAIL_FOOTER = (
    '<div style="margin-top:32px;padding-top:16px;border-top:1px solid #e2e8f0;text-align:center">'
    '<p style="font-size:.75rem;color:#94a3b8;margin:0">'
    'This is an automated email from SkillBaseHire. Please do not reply.'
    '</p></div>'
)


def _email_wrap(content_html):
    """Wrap email content in branded SkillBaseHire container."""
    return (
        '<div style="font-family:Inter,sans-serif;max-width:520px;margin:0 auto;'
        'padding:32px 24px;background:#fff;border-radius:12px">'
        '<div style="text-align:center;margin-bottom:28px">'
        '<span style="font-size:1.25rem;font-weight:800;color:#0f172a;letter-spacing:-.3px">'
        'SkillBaseHire</span></div>'
        + content_html
        + _EMAIL_FOOTER
        + '</div>'
    )


def _send_email(to_email, subject, html_body):
    """Central email dispatcher. Returns True on success, False on failure."""
    smtp_user = os.environ.get('SMTP_USER', '').strip()
    smtp_pass = os.environ.get('SMTP_PASS', '').strip()
    smtp_host = os.environ.get('SMTP_HOST', 'smtp.gmail.com').strip()
    smtp_port = int(os.environ.get('SMTP_PORT', 587))
    # Gmail requires the envelope sender to match the authenticated account.
    # Use smtp_user as envelope sender; display name/address is set in From header.
    envelope_from = smtp_user or _SMTP_FROM_EMAIL
    from_header   = f'{_SMTP_FROM_NAME} <{envelope_from}>'

    missing = [v for v, val in [('SMTP_USER', smtp_user), ('SMTP_PASS', smtp_pass)] if not val]
    if missing:
        app.logger.warning(
            '[EMAIL_SMTP] Missing env vars: %s — email "%s" to %s NOT sent. '
            'Set SMTP_USER, SMTP_PASS (and optionally SMTP_HOST, SMTP_PORT) in environment.',
            ', '.join(missing), subject, to_email
        )
        return False

    msg = MIMEMultipart('alternative')
    msg['Subject'] = subject
    msg['From']    = from_header
    msg['To']      = to_email
    msg.attach(MIMEText(html_body, 'html'))

    app.logger.info('[EMAIL_SMTP] Connecting to %s:%s as %s', smtp_host, smtp_port, smtp_user)
    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=15) as server:
            server.ehlo()
            server.starttls()
            server.ehlo()
            server.login(smtp_user, smtp_pass)
            server.sendmail(envelope_from, to_email, msg.as_string())
        app.logger.info('[EMAIL] sent "%s" to %s', subject, to_email)
        return True
    except smtplib.SMTPAuthenticationError as exc:
        app.logger.error('[EMAIL_ERROR] SMTP authentication failed for %s — wrong password or App Password not used: %s', smtp_user, exc)
        return False
    except Exception as exc:
        app.logger.error('[EMAIL_ERROR] failed "%s" to %s via %s:%s — %s',
                         subject, to_email, smtp_host, smtp_port, exc)
        return False


def _send_security_alert(attempted_email, ip, user_agent, page):
    now     = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')
    subject = 'Security Alert: Invalid Manage Questions Token Attempt'
    content = (
        '<div style="background:#fef2f2;border:1px solid #fecaca;border-radius:10px;'
        'padding:20px 24px;margin-bottom:20px">'
        '<h2 style="font-size:1.1rem;font-weight:700;color:#dc2626;margin:0 0 4px">Security Alert</h2>'
        '<p style="color:#7f1d1d;margin:0;font-size:.9rem">An invalid admin token was used on SkillBaseHire.</p>'
        '</div>'
        '<table style="width:100%;border-collapse:collapse;font-size:.875rem;color:#334155">'
        f'<tr><td style="padding:8px 0;border-bottom:1px solid #f1f5f9;font-weight:600;width:160px">Attempted by</td><td style="padding:8px 0;border-bottom:1px solid #f1f5f9">{attempted_email}</td></tr>'
        f'<tr><td style="padding:8px 0;border-bottom:1px solid #f1f5f9;font-weight:600">Date &amp; Time</td><td style="padding:8px 0;border-bottom:1px solid #f1f5f9">{now}</td></tr>'
        f'<tr><td style="padding:8px 0;border-bottom:1px solid #f1f5f9;font-weight:600">IP Address</td><td style="padding:8px 0;border-bottom:1px solid #f1f5f9">{ip or "Unknown"}</td></tr>'
        f'<tr><td style="padding:8px 0;border-bottom:1px solid #f1f5f9;font-weight:600">Browser/Device</td><td style="padding:8px 0;border-bottom:1px solid #f1f5f9;word-break:break-all">{user_agent or "Unknown"}</td></tr>'
        f'<tr><td style="padding:8px 0;font-weight:600">Page Attempted</td><td style="padding:8px 0">{page}</td></tr>'
        '</table>'
    )
    smtp_user = os.environ.get('SMTP_USER')
    smtp_pass = os.environ.get('SMTP_PASS')
    if not smtp_user or not smtp_pass:
        app.logger.warning('[SECURITY_ALERT] SMTP not configured — alert not sent. %s attempted %s at %s from %s',
                           attempted_email, page, now, ip)
        return
    _send_email(_SUPER_ADMIN_EMAIL, subject, _email_wrap(content))


def get_current_user():
    if 'user_id' not in session:
        return None
    return get_db().execute('SELECT * FROM users WHERE id = ?', [session['user_id']]).fetchone()


_PW_SPECIAL = r'[@#$%&*!]'

def validate_password(password):
    if not password:
        return 'Password is required'
    if password != password.strip():
        return 'Password cannot start or end with spaces'
    if len(password) < 8:
        return 'Password must be at least 8 characters'
    if not re.search(r'[A-Z]', password):
        return 'Password must include at least one uppercase letter'
    if not re.search(r'[a-z]', password):
        return 'Password must include at least one lowercase letter'
    if not re.search(r'\d', password):
        return 'Password must include at least one number'
    if not re.search(_PW_SPECIAL, password):
        return 'Password must include at least one special character (@#$%&*!)'
    return None


def allowed_resume(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_RESUME_EXT


def _gen_email_token():
    """Return (raw_token, token_hash, expires_at). Store only the hash in DB."""
    raw = secrets.token_urlsafe(32)
    h = hashlib.sha256(raw.encode()).hexdigest()
    expires = datetime.utcnow() + timedelta(hours=24)
    return raw, h, expires


def send_verification_email(to_email, name, raw_token):
    verify_url = url_for('verify_email', token=raw_token, _external=True)
    subject = 'Verify your SkillBaseHire account'
    content = (
        f'<h2 style="font-size:1.25rem;font-weight:700;color:#1e293b;margin:0 0 8px">Hi {name},</h2>'
        f'<p style="color:#475569;margin:0 0 24px">Click the button below to verify your email address and activate your account.</p>'
        f'<div style="text-align:center;margin:28px 0">'
        f'<a href="{verify_url}" style="display:inline-block;background:#2563eb;color:#fff;text-decoration:none;font-weight:600;padding:12px 32px;border-radius:8px">Verify Email Address</a>'
        f'</div>'
        f'<p style="color:#94a3b8;font-size:.8125rem;margin:0">This link expires in 24 hours. If you didn\'t create an account, you can ignore this email.</p>'
    )
    sent = _send_email(to_email, subject, _email_wrap(content))
    if not sent:
        app.logger.info('[EMAIL_DEV] Verification link for %s: %s', to_email, verify_url)


def send_password_reset_email(to_email, name, reset_url):
    subject = 'Reset your SkillBaseHire password'
    content = (
        f'<p style="color:#1e293b;margin:0 0 8px">Hi,</p>'
        f'<p style="color:#475569;margin:0 0 24px">We received a request to reset your SkillBaseHire password.</p>'
        f'<p style="color:#475569;margin:0 0 16px">Click below to reset your password:</p>'
        f'<div style="text-align:center;margin:28px 0">'
        f'<a href="{reset_url}" style="display:inline-block;background:#2563eb;color:#fff;text-decoration:none;font-weight:600;padding:12px 32px;border-radius:8px">Reset Password</a>'
        f'</div>'
        f'<p style="color:#94a3b8;font-size:.8125rem;margin:0 0 8px">This link will expire in 15 minutes.</p>'
        f'<p style="color:#94a3b8;font-size:.8125rem;margin:0">If you did not request this, please ignore this email.</p>'
    )
    sent = _send_email(to_email, subject, _email_wrap(content))
    if not sent:
        if _is_staging:
            app.logger.info('[FORGOT_PW_STAGING] SMTP not configured. Password reset link generated but email not sent. Link: %s', reset_url)
        else:
            app.logger.warning('[FORGOT_PW] SMTP not configured. Password reset link generated but email not sent.')


def parse_skills_data(raw):
    """Parse 'Name:verified,Name:verified' string into list of dicts."""
    if not raw:
        return []
    result = []
    for item in raw.split(','):
        parts = item.strip().split(':')
        if len(parts) == 2:
            result.append({'name': parts[0], 'verified': parts[1] == '1'})
    return result


# ── Profile photo serve (DB-backed, survives ephemeral filesystem) ────────────

_PHOTO_MIME = {'jpg': 'image/jpeg', 'jpeg': 'image/jpeg', 'png': 'image/png', 'webp': 'image/webp'}

@app.route('/photo/<path:filename>')
def serve_profile_photo(filename):
    db = get_db()
    row = None
    try:
        if filename.startswith('candidate_'):
            row = db.execute(
                'SELECT profile_photo_data, profile_photo_mime FROM candidate_profiles WHERE profile_photo=?',
                [filename]).fetchone()
        elif filename.startswith('recruiter_'):
            row = db.execute(
                'SELECT profile_photo_data, profile_photo_mime FROM recruiter_profiles WHERE profile_photo=?',
                [filename]).fetchone()
    except Exception:
        pass
    if row and row.get('profile_photo_data'):
        raw = row['profile_photo_data']
        if isinstance(raw, memoryview):
            raw = raw.tobytes().decode('ascii')
        mime = row.get('profile_photo_mime') or 'image/jpeg'
        return send_file(io.BytesIO(base64.b64decode(raw)), mimetype=mime)
    # Fallback: disk (for photos uploaded before DB migration)
    disk_path = os.path.join(PHOTO_UPLOAD_FOLDER, filename)
    if os.path.exists(disk_path):
        return send_from_directory(PHOTO_UPLOAD_FOLDER, filename)
    return ('', 404)


# ── Home ─────────────────────────────────────────────────────────────────────

@app.route('/')
def home():
    db = get_db()
    recent_jobs = db.execute('''
        SELECT j.*, rp.company AS company_name,
               (SELECT GROUP_CONCAT(s.name, ', ')
                FROM job_skills js JOIN skills s ON js.skill_id = s.id
                WHERE js.job_id = j.id) AS skills_list
        FROM jobs j
        JOIN recruiter_profiles rp ON j.recruiter_id = rp.user_id
        WHERE j.active = 1
        ORDER BY j.created_at DESC LIMIT 6
    ''').fetchall()

    stats = {
        'jobs': db.execute('SELECT COUNT(*) FROM jobs WHERE active=1').fetchone()[0],
        'candidates': db.execute("SELECT COUNT(*) FROM users WHERE role='candidate'").fetchone()[0],
        'companies': db.execute("SELECT COUNT(DISTINCT company) FROM jobs WHERE active=1").fetchone()[0],
    }
    return render_template('index.html', jobs=recent_jobs, stats=stats, user=get_current_user())


# ── Jobs ─────────────────────────────────────────────────────────────────────

@app.route('/jobs')
def jobs():
    db = get_db()
    search   = request.args.get('q', '').strip()
    location = request.args.get('location', '').strip()
    job_type = request.args.get('type', '').strip()

    query = '''
        SELECT j.*, rp.company AS company_name,
               (SELECT GROUP_CONCAT(s.name, ', ')
                FROM job_skills js JOIN skills s ON js.skill_id = s.id
                WHERE js.job_id = j.id) AS skills_list,
               (SELECT GROUP_CONCAT(js.skill_id, ',')
                FROM job_skills js
                WHERE js.job_id = j.id) AS skill_ids_list
        FROM jobs j
        JOIN recruiter_profiles rp ON j.recruiter_id = rp.user_id
        WHERE j.active = 1
    '''
    params = []
    if search:
        query += ' AND (j.title LIKE ? OR j.description LIKE ? OR j.company LIKE ?)'
        params += [f'%{search}%', f'%{search}%', f'%{search}%']
    if location:
        query += ' AND j.location LIKE ?'
        params.append(f'%{location}%')
    if job_type:
        query += ' AND j.job_type = ?'
        params.append(job_type)
    query += ' ORDER BY j.created_at DESC'

    job_rows = db.execute(query, params).fetchall()

    # Candidate profile data for match calculation and modal pre-fill
    verified_skills = set()
    applied_job_ids = set()
    resume_filename = None
    if session.get('role') == 'candidate' and session.get('user_id'):
        rows = db.execute('''
            SELECT s.name FROM user_skills us
            JOIN skills s ON us.skill_id = s.id
            WHERE us.user_id = ? AND us.verified = 1
        ''', [session['user_id']]).fetchall()
        verified_skills = {r['name'] for r in rows}
        applied_rows = db.execute(
            'SELECT job_id FROM applications WHERE candidate_id=?',
            [session['user_id']]
        ).fetchall()
        applied_job_ids = {r['job_id'] for r in applied_rows}
        prof_row = db.execute(
            'SELECT resume_filename FROM candidate_profiles WHERE user_id=?',
            [session['user_id']]
        ).fetchone()
        if prof_row:
            resume_filename = prof_row['resume_filename']

    colors = ['#635BFF','#FF5C35','#3B82F6','#10B981','#F59E0B',
              '#8B5CF6','#0EA5E9','#14B8A6','#F97316','#EF4444']

    jobs_data = []
    for i, job in enumerate(job_rows):
        skills      = [s.strip() for s in job['skills_list'].split(', ')] if job['skills_list'] else []
        skill_ids   = [int(x) for x in job['skill_ids_list'].split(',')] if job['skill_ids_list'] else []
        skill_id_map = {name: sid for name, sid in zip(skills, skill_ids)}
        company     = job['company_name'] or job['company']
        color       = colors[i % len(colors)]
        verified_in = [s for s in skills if s in verified_skills]
        missing_in  = [{'name': s, 'id': skill_id_map.get(s)} for s in skills if s not in verified_skills]
        match_pct   = round(len(verified_in) / len(skills) * 100) if skills else 0

        try:
            posted_dt = datetime.strptime(str(job['created_at'])[:19], '%Y-%m-%d %H:%M:%S')
            days_ago  = (datetime.utcnow() - posted_dt).days
            posted_str = 'Today' if days_ago == 0 else f'{days_ago}d ago'
        except Exception:
            days_ago, posted_str = 0, 'Recently'

        signals = []
        if 'remote' in job['location'].lower():
            signals.append({'label': 'Remote', 'color': 'teal'})
        if days_ago <= 2:
            signals.append({'label': 'New', 'color': 'green'})
        if job['job_type'] == 'Contract':
            signals.append({'label': 'Contract', 'color': 'orange'})

        sal_min = (job['salary_min'] // 100000) if job['salary_min'] else 0
        sal_max = (job['salary_max'] // 100000) if job['salary_max'] else 0

        jobs_data.append({
            'id':         job['id'],
            'title':      job['title'],
            'company':    company,
            'mark':       company[0].upper(),
            'color':      color,
            'location':   job['location'],
            'type':       job['job_type'],
            'expRange':   [0, 5],
            'salRange':   [sal_min, sal_max],
            'posted':     posted_str,
            'postedSort': days_ago,
            'match':      match_pct,
            'signals':    signals,
            'reqSkills':  skills,
            'verifiedIn': verified_in,
            'missingIn':  missing_in,
            'skillTestMandatory': bool(job['skill_test_mandatory']),
            'desc':       job['description'] or '',
            'requirements': [r.strip() for r in (job['requirements'] or '').split('\n') if r.strip()],
            'perks':      [],
            'saved':      False,
            'applied':    job['id'] in applied_job_ids,
        })

    return render_template('jobs.html',
                           jobs_data=jobs_data,
                           search=search, location=location, job_type=job_type,
                           user=get_current_user(),
                           resume_filename=resume_filename)


@app.route('/jobs/<int:job_id>')
def job_detail(job_id):
    db = get_db()
    job = db.execute('''
        SELECT j.*, rp.company AS company_name, rp.company_bio, rp.website,
               u.name AS recruiter_name
        FROM jobs j
        JOIN recruiter_profiles rp ON j.recruiter_id = rp.user_id
        JOIN users u ON j.recruiter_id = u.id
        WHERE j.id = ? AND j.active = 1
    ''', [job_id]).fetchone()

    if not job:
        flash('Job not found.', 'error')
        return redirect(url_for('jobs'))

    job_skills = db.execute('''
        SELECT s.* FROM job_skills js JOIN skills s ON js.skill_id = s.id
        WHERE js.job_id = ?
    ''', [job_id]).fetchall()

    already_applied = False
    my_skill_names = set()
    my_verified_names = set()
    if session.get('role') == 'candidate':
        already_applied = db.execute(
            'SELECT id FROM applications WHERE job_id=? AND candidate_id=?',
            [job_id, session['user_id']]
        ).fetchone() is not None
        for row in db.execute('''
            SELECT s.name, us.verified FROM user_skills us
            JOIN skills s ON us.skill_id = s.id WHERE us.user_id = ?
        ''', [session['user_id']]).fetchall():
            my_skill_names.add(row['name'])
            if row['verified']:
                my_verified_names.add(row['name'])

    return render_template('job_detail.html', job=job, job_skills=job_skills,
                           already_applied=already_applied,
                           my_skill_names=my_skill_names,
                           my_verified_names=my_verified_names,
                           user=get_current_user())


# ── Skills page ───────────────────────────────────────────────────────────────

@app.route('/skills')
def skills_page():
    db = get_db()
    all_skills = db.execute('SELECT * FROM skills ORDER BY category, name').fetchall()

    user_added = set()
    user_verified = set()
    if session.get('role') == 'candidate':
        for row in db.execute(
            'SELECT skill_id, verified FROM user_skills WHERE user_id=?',
            [session['user_id']]
        ).fetchall():
            user_added.add(row['skill_id'])
            if row['verified']:
                user_verified.add(row['skill_id'])

    skills_by_category = {}
    for s in all_skills:
        cat = s['category']
        if cat not in skills_by_category:
            skills_by_category[cat] = []
        skills_by_category[cat].append({
            'id': s['id'], 'name': s['name'], 'description': s['description'],
            'added': s['id'] in user_added, 'verified': s['id'] in user_verified,
        })

    return render_template('skills.html', skills_by_category=skills_by_category,
                           user=get_current_user())


@app.route('/skills/add/<int:skill_id>', methods=['POST'])
@candidate_required
def add_skill(skill_id):
    db = get_db()
    db.execute('INSERT OR IGNORE INTO user_skills (user_id, skill_id) VALUES (?, ?)',
               [session['user_id'], skill_id])
    db.commit()
    flash('Skill added to your profile!', 'success')
    return redirect(url_for('skills_page'))


QUIZ_BANK = {
    'Python': [
        {'q': 'What is the output of type([])?', 'options': ["<class 'list'>", "<class 'array'>", "<class 'tuple'>", "list"], 'answer': 0},
        {'q': 'Which keyword defines a function in Python?', 'options': ['func', 'def', 'function', 'lambda'], 'answer': 1},
        {'q': 'What does PEP stand for?', 'options': ['Python Enhancement Proposal', 'Python Execution Protocol', 'Program Enhancement Package', 'Python Error Package'], 'answer': 0},
        {'q': 'How do you create a list comprehension?', 'options': ['[x for x in range(5)]', '{x for x in range(5)}', '(x for x in range(5))', 'list(x for x in 5)'], 'answer': 0},
        {'q': 'Which method removes and returns the last element of a list?', 'options': ['remove()', 'pop()', 'delete()', 'discard()'], 'answer': 1},
        {'q': 'What is a Python decorator?', 'options': ['A class inheritance pattern', 'A function that modifies another function', 'A type of loop', 'A module import style'], 'answer': 1},
    ],
    'JavaScript': [
        {'q': 'Which operator checks strict equality?', 'options': ['==', '===', '=', '!='], 'answer': 1},
        {'q': 'What does typeof null return?', 'options': ['null', 'undefined', 'object', 'boolean'], 'answer': 2},
        {'q': 'Which method adds an element to the end of an array?', 'options': ['push()', 'pop()', 'shift()', 'append()'], 'answer': 0},
        {'q': 'What is a closure in JavaScript?', 'options': ['A function with access to its outer scope', 'A class constructor', 'A type of loop', 'An import statement'], 'answer': 0},
        {'q': 'What does the spread operator (...) do?', 'options': ['Deletes array elements', 'Expands iterable elements', 'Creates a new loop', 'Declares a variable'], 'answer': 1},
        {'q': 'Which keyword creates a block-scoped variable?', 'options': ['var', 'let', 'const', 'Both let and const'], 'answer': 3},
    ],
    'React': [
        {'q': 'Which hook handles side effects?', 'options': ['useState', 'useEffect', 'useContext', 'useReducer'], 'answer': 1},
        {'q': 'What is JSX?', 'options': ['A database query language', 'A syntax extension for JavaScript', 'A CSS preprocessor', 'A testing framework'], 'answer': 1},
        {'q': 'What does useState return?', 'options': ['A value only', 'A setter only', 'A value and a setter function', 'An event handler'], 'answer': 2},
        {'q': 'Which lifecycle runs after component mounts?', 'options': ['componentWillMount', 'componentDidMount', 'componentDidUpdate', 'render'], 'answer': 1},
        {'q': 'What is the purpose of the key prop in lists?', 'options': ['Encrypts data', 'Helps React identify changed elements', 'Declares a CSS class', 'Triggers re-renders'], 'answer': 1},
        {'q': 'Which hook provides access to context?', 'options': ['useRef', 'useMemo', 'useContext', 'useCallback'], 'answer': 2},
    ],
    'SQL': [
        {'q': 'Which clause filters rows?', 'options': ['ORDER BY', 'GROUP BY', 'WHERE', 'HAVING'], 'answer': 2},
        {'q': 'What does JOIN do?', 'options': ['Combines rows from two tables', 'Deletes rows', 'Adds a column', 'Sorts results'], 'answer': 0},
        {'q': 'Which aggregate function counts rows?', 'options': ['SUM()', 'COUNT()', 'MAX()', 'AVG()'], 'answer': 1},
        {'q': 'What is a PRIMARY KEY?', 'options': ['A unique identifier for each row', 'The first column', 'An index', 'A foreign reference'], 'answer': 0},
        {'q': 'Which statement removes all rows from a table without deleting the table?', 'options': ['DELETE', 'DROP', 'TRUNCATE', 'REMOVE'], 'answer': 2},
        {'q': 'What does DISTINCT do in a SELECT statement?', 'options': ['Sorts results', 'Removes duplicate rows', 'Filters by condition', 'Joins tables'], 'answer': 1},
    ],
    'Java': [
        {'q': 'Which keyword is used to define a class in Java?', 'options': ['class', 'Class', 'define', 'object'], 'answer': 0},
        {'q': 'What is the default value of an int variable in Java?', 'options': ['null', '0', '1', 'undefined'], 'answer': 1},
        {'q': 'Which collection interface allows duplicate elements?', 'options': ['Set', 'Map', 'List', 'SortedSet'], 'answer': 2},
        {'q': 'What does JVM stand for?', 'options': ['Java Variable Machine', 'Java Virtual Machine', 'Java Verified Method', 'Java Visual Manager'], 'answer': 1},
        {'q': 'Which access modifier makes a member accessible only within its own class?', 'options': ['public', 'protected', 'private', 'package-private'], 'answer': 2},
        {'q': 'What is method overloading in Java?', 'options': ['Same method name, different parameters', 'Same method name, different class', 'Using abstract classes', 'Extending a parent class method'], 'answer': 0},
        {'q': 'Which keyword prevents a method from being overridden?', 'options': ['static', 'abstract', 'final', 'private'], 'answer': 2},
        {'q': 'What is the parent class of all Java classes?', 'options': ['Base', 'Object', 'Class', 'Super'], 'answer': 1},
    ],
    'Selenium': [
        {'q': 'Which method locates a web element by its ID?', 'options': ['driver.findElement(By.name())', 'driver.findElement(By.id())', 'driver.getElement()', 'driver.locateById()'], 'answer': 1},
        {'q': 'What does WebDriver.get() do?', 'options': ['Gets the page title', 'Opens a URL in the browser', 'Returns the current URL', 'Gets the page source'], 'answer': 1},
        {'q': 'Which wait polls until a condition is met with configurable polling interval?', 'options': ['ImplicitWait', 'ExplicitWait', 'FluentWait', 'StaticWait'], 'answer': 2},
        {'q': 'What is the correct way to click a button in Selenium?', 'options': ['element.press()', 'element.click()', 'element.trigger()', 'driver.click(element)'], 'answer': 1},
        {'q': 'Which method retrieves visible text of an element?', 'options': ['element.getText()', 'element.getValue()', 'element.getContent()', 'element.innerHTML()'], 'answer': 0},
        {'q': 'How do you switch to an iframe in Selenium?', 'options': ['driver.switchTo().frame()', 'driver.selectFrame()', 'driver.openFrame()', 'driver.navigateFrame()'], 'answer': 0},
        {'q': 'Which XPath selects all input elements?', 'options': ['/input', '//input', './input', 'input[]'], 'answer': 1},
        {'q': 'What does driver.quit() do differently from driver.close()?', 'options': ['Same behavior', 'Quits all windows and ends session', 'Closes only the active tab', 'Refreshes the browser'], 'answer': 1},
    ],
    'Playwright': [
        {'q': 'Which Playwright method navigates to a URL?', 'options': ['page.visit()', 'page.goto()', 'page.open()', 'page.navigate()'], 'answer': 1},
        {'q': 'How do you click an element in Playwright?', 'options': ['page.press()', 'page.tap()', 'page.click()', 'page.select()'], 'answer': 2},
        {'q': 'Which method waits for a selector to appear?', 'options': ['page.awaitSelector()', 'page.waitForSelector()', 'page.findElement()', 'page.expectElement()'], 'answer': 1},
        {'q': 'What does page.fill() do in Playwright?', 'options': ['Clears a form', 'Types text into an input field', 'Submits a form', 'Validates form data'], 'answer': 1},
        {'q': 'Which Playwright method takes a screenshot?', 'options': ['page.capture()', 'page.screen()', 'page.screenshot()', 'page.snap()'], 'answer': 2},
        {'q': 'How do you handle a dialog/alert in Playwright?', 'options': ['page.dismissDialog()', 'page.on("dialog", handler)', 'page.acceptAlert()', 'page.dismissAlert()'], 'answer': 1},
        {'q': 'Which command runs Playwright tests in headed mode?', 'options': ['npx playwright test --visible', 'npx playwright test --headed', 'npx playwright test --gui', 'npx playwright test --open'], 'answer': 1},
        {'q': 'What is a Playwright fixture?', 'options': ['A test helper for common setup/teardown', 'A CSS selector strategy', 'A browser configuration file', 'A type of assertion'], 'answer': 0},
    ],
    'Appium': [
        {'q': 'What type of applications can Appium test?', 'options': ['Only Android', 'Only iOS', 'Both mobile and desktop native apps', 'Only web apps'], 'answer': 2},
        {'q': 'Which capability sets the application package in Android Appium tests?', 'options': ['app', 'appPackage', 'appBundle', 'appActivity'], 'answer': 1},
        {'q': 'What is the Appium server used for?', 'options': ['Running tests on the browser', 'Bridging test scripts and mobile devices', 'Compiling test code', 'Managing test data'], 'answer': 1},
        {'q': 'Which locator strategy finds elements by accessibility ID?', 'options': ['By.id()', 'MobileBy.AccessibilityId()', 'By.accessibilityId()', 'MobileBy.id()'], 'answer': 1},
        {'q': 'What does the desired capability "platformName" specify?', 'options': ['The device model', 'The OS (Android/iOS)', 'The app version', 'The automation engine'], 'answer': 1},
        {'q': 'Which Appium driver is used for iOS automation?', 'options': ['UIAutomator2', 'XCUITest', 'Espresso', 'Instrumentation'], 'answer': 1},
        {'q': 'How do you perform a swipe gesture in Appium?', 'options': ['driver.swipe()', 'new TouchAction(driver).press().moveTo().release().perform()', 'driver.gesture("swipe")', 'driver.scroll()'], 'answer': 1},
        {'q': 'What is Appium Inspector used for?', 'options': ['Running test scripts', 'Inspecting UI elements and their properties', 'Generating test reports', 'Managing device connections'], 'answer': 1},
    ],
    'API Testing': [
        {'q': 'Which HTTP method is used to retrieve data from a server?', 'options': ['POST', 'PUT', 'GET', 'DELETE'], 'answer': 2},
        {'q': 'What HTTP status code indicates a successful resource creation?', 'options': ['200', '201', '204', '400'], 'answer': 1},
        {'q': 'What does a 404 status code mean?', 'options': ['Internal server error', 'Unauthorized', 'Resource not found', 'Bad request'], 'answer': 2},
        {'q': 'Which tool is commonly used for API testing?', 'options': ['Selenium', 'Postman', 'JUnit', 'Appium'], 'answer': 1},
        {'q': 'What is the purpose of an API authentication token?', 'options': ['To compress data', 'To verify the identity of the requester', 'To format JSON responses', 'To cache API responses'], 'answer': 1},
        {'q': 'What does REST stand for?', 'options': ['Remote Execution Service Technology', 'Representational State Transfer', 'Reliable Endpoint Service Transfer', 'Resource Execution Standard Test'], 'answer': 1},
        {'q': 'Which HTTP method updates an existing resource completely?', 'options': ['GET', 'POST', 'PUT', 'PATCH'], 'answer': 2},
        {'q': 'What is JSON?', 'options': ['A database type', 'A JavaScript framework', 'A lightweight data interchange format', 'A testing protocol'], 'answer': 2},
    ],
    'Manual Testing': [
        {'q': 'What is a test case?', 'options': ['A bug report', 'A set of conditions to verify a feature', 'A deployment script', 'A performance benchmark'], 'answer': 1},
        {'q': 'What is regression testing?', 'options': ['Testing new features only', 'Re-testing after changes to ensure nothing broke', 'Testing performance under load', 'Initial feature testing'], 'answer': 1},
        {'q': 'What is the difference between severity and priority in bug reporting?', 'options': ['They are the same thing', 'Severity is impact on system; priority is urgency of fix', 'Priority is impact on system; severity is urgency', 'Both refer to bug frequency'], 'answer': 1},
        {'q': 'What is exploratory testing?', 'options': ['Testing from a test script', 'Simultaneous learning, test design, and execution', 'Automated regression testing', 'User acceptance testing'], 'answer': 1},
        {'q': 'What is a test plan?', 'options': ['A single test case', 'A document describing testing strategy, scope, and resources', 'A bug tracking spreadsheet', 'An automated test script'], 'answer': 1},
        {'q': 'What is boundary value analysis?', 'options': ['Testing random inputs', 'Testing at the edges of valid input ranges', 'Checking UI boundaries', 'Testing network limits'], 'answer': 1},
        {'q': 'Which testing type validates the system meets business requirements?', 'options': ['Unit testing', 'Integration testing', 'User Acceptance Testing (UAT)', 'Performance testing'], 'answer': 2},
        {'q': 'What is a defect lifecycle?', 'options': ['Time to close a defect', 'Stages a bug goes through from discovery to closure', 'Number of defects per sprint', 'Defect density metric'], 'answer': 1},
    ],
}

DEFAULT_QUESTIONS = lambda name: [
    {'q': f'Have you used {name} in a real project?', 'options': ['Yes, extensively', 'Yes, moderately', 'Yes, briefly', 'No, only studied it'], 'answer': 0},
    {'q': f'How would you rate your {name} expertise?', 'options': ['Expert (5+ years)', 'Advanced (3-5 years)', 'Intermediate (1-3 years)', 'Beginner (<1 year)'], 'answer': 0},
    {'q': f'Can you explain core {name} concepts to others?', 'options': ['Yes, confidently', 'Yes, mostly', 'Partially', 'Not yet'], 'answer': 0},
    {'q': f'Do you keep up with {name} updates and best practices?', 'options': ['Yes, actively', 'Mostly yes', 'Occasionally', 'Rarely'], 'answer': 0},
]

# Exam metadata per skill
SKILL_META = {
    'Java':           {'duration': 45, 'questions': 8, 'passing': 70, 'attempts': 2, 'level': 'Intermediate'},
    'Selenium':       {'duration': 45, 'questions': 8, 'passing': 70, 'attempts': 2, 'level': 'Intermediate'},
    'Playwright':     {'duration': 45, 'questions': 8, 'passing': 70, 'attempts': 2, 'level': 'Intermediate'},
    'API Testing':    {'duration': 40, 'questions': 8, 'passing': 70, 'attempts': 2, 'level': 'Intermediate'},
    'Manual Testing': {'duration': 35, 'questions': 8, 'passing': 70, 'attempts': 2, 'level': 'Beginner–Intermediate'},
    'Appium':         {'duration': 45, 'questions': 8, 'passing': 70, 'attempts': 2, 'level': 'Intermediate'},
    'SQL':            {'duration': 35, 'questions': 6, 'passing': 70, 'attempts': 2, 'level': 'Intermediate'},
    'Python':         {'duration': 30, 'questions': 6, 'passing': 66, 'attempts': 3, 'level': 'Intermediate'},
    'JavaScript':     {'duration': 30, 'questions': 6, 'passing': 66, 'attempts': 3, 'level': 'Intermediate'},
    'React':          {'duration': 30, 'questions': 6, 'passing': 66, 'attempts': 3, 'level': 'Intermediate'},
}
DEFAULT_META = {'duration': 25, 'questions': 4, 'passing': 66, 'attempts': 3, 'level': 'Intermediate'}


def _db_q_to_quiz(row):
    """Convert a skill_questions DB row to QUIZ_BANK format."""
    options = [row['option_a'], row['option_b'], row['option_c'], row['option_d']]
    answer = ord(row['correct_option'].upper()) - ord('A')
    return {'q': row['question_text'], 'options': options, 'answer': answer}


@app.route('/skills/verify/<int:skill_id>/instructions')
@candidate_required
def exam_instructions(skill_id):
    db = get_db()
    skill = db.execute('SELECT * FROM skills WHERE id=?', [skill_id]).fetchone()
    if not skill:
        flash('Skill not found.', 'error')
        return redirect(url_for('skills_page'))
    meta = dict(SKILL_META.get(skill['name'], DEFAULT_META))
    db_q_count = db.execute(
        'SELECT COUNT(*) FROM skill_questions WHERE skill_id=?', [skill_id]
    ).fetchone()[0]
    if db_q_count:
        meta['questions'] = db_q_count
    existing = db.execute(
        'SELECT score, verified FROM user_skills WHERE user_id=? AND skill_id=?',
        [session['user_id'], skill_id]
    ).fetchone()
    return render_template('exam_instructions.html', skill=skill, meta=meta,
                           existing=existing, user=get_current_user())


@app.route('/skills/verify/<int:skill_id>', methods=['GET', 'POST'])
@candidate_required
def verify_skill(skill_id):
    db = get_db()
    skill = db.execute('SELECT * FROM skills WHERE id=?', [skill_id]).fetchone()
    if not skill:
        flash('Skill not found.', 'error')
        return redirect(url_for('skills_page'))

    db_rows = db.execute(
        'SELECT * FROM skill_questions WHERE skill_id=? ORDER BY id', [skill_id]
    ).fetchall()
    if db_rows:
        questions = [_db_q_to_quiz(r) for r in db_rows]
        meta = dict(SKILL_META.get(skill['name'], DEFAULT_META))
        meta['questions'] = len(questions)
    else:
        questions = QUIZ_BANK.get(skill['name'], DEFAULT_QUESTIONS(skill['name']))
        meta = SKILL_META.get(skill['name'], DEFAULT_META)

    if request.method == 'POST':
        integrity_ended = request.form.get('integrity_ended') == '1'
        fs_exit_count = request.form.get('fs_exit_count', 0, type=int)

        if integrity_ended:
            existing = db.execute(
                'SELECT id FROM user_skills WHERE user_id=? AND skill_id=?',
                [session['user_id'], skill_id]
            ).fetchone()
            if existing:
                db.execute('UPDATE user_skills SET verified=0, score=0 WHERE user_id=? AND skill_id=?',
                           [session['user_id'], skill_id])
            else:
                db.execute('INSERT INTO user_skills (user_id, skill_id, verified, score) VALUES (?,?,0,0)',
                           [session['user_id'], skill_id])
            db.commit()
            session['exam_ended_data'] = {
                'skill_id': skill_id,
                'skill_name': skill['name'],
                'fs_exit_count': fs_exit_count,
            }
            return redirect(url_for('exam_ended', skill_id=skill_id))

        score = sum(
            1 for i, q in enumerate(questions)
            if request.form.get(f'q{i}', type=int) == q['answer']
        )
        total = len(questions)
        pct = int((score / total) * 100)
        verified = 1 if pct >= meta['passing'] else 0

        time_elapsed = request.form.get('time_elapsed', 0, type=int)
        mins = time_elapsed // 60
        secs = time_elapsed % 60
        time_taken_str = f'{mins}:{secs:02d}'

        existing = db.execute(
            'SELECT id FROM user_skills WHERE user_id=? AND skill_id=?',
            [session['user_id'], skill_id]
        ).fetchone()
        if existing:
            db.execute(
                'UPDATE user_skills SET verified=?, score=?, correct_answers=?, time_taken_secs=? WHERE user_id=? AND skill_id=?',
                [verified, pct, score, time_elapsed, session['user_id'], skill_id]
            )
        else:
            db.execute(
                'INSERT INTO user_skills (user_id, skill_id, verified, score, correct_answers, time_taken_secs) VALUES (?,?,?,?,?,?)',
                [session['user_id'], skill_id, verified, pct, score, time_elapsed]
            )
        db.commit()

        session['exam_result'] = {
            'skill_id': skill_id,
            'skill_name': skill['name'],
            'score': pct,
            'correct': score,
            'total': total,
            'verified': bool(verified),
            'passing': meta['passing'],
            'duration': meta['duration'],
            'time_taken': time_taken_str,
        }
        return redirect(url_for('exam_result', skill_id=skill_id))

    # GET — redirect to result if already terminated this session
    if session.get('exam_result', {}).get('skill_id') == skill_id:
        return redirect(url_for('exam_result', skill_id=skill_id))

    # Show active exam; add skill to profile if not already present
    db.execute('INSERT OR IGNORE INTO user_skills (user_id, skill_id) VALUES (?,?)',
               [session['user_id'], skill_id])
    db.commit()
    return render_template('verify_skill.html', skill=skill, questions=questions,
                           meta=meta, user=get_current_user())


@app.route('/skills/verify/<int:skill_id>/auto-submit', methods=['POST'])
@candidate_required
def exam_auto_submit(skill_id):
    db = get_db()
    skill = db.execute('SELECT * FROM skills WHERE id=?', [skill_id]).fetchone()
    if not skill:
        return jsonify({'ok': False, 'error': 'Skill not found'}), 404

    db_rows = db.execute(
        'SELECT * FROM skill_questions WHERE skill_id=? ORDER BY id', [skill_id]
    ).fetchall()
    if db_rows:
        questions = [_db_q_to_quiz(r) for r in db_rows]
        meta = dict(SKILL_META.get(skill['name'], DEFAULT_META))
        meta['questions'] = len(questions)
    else:
        questions = QUIZ_BANK.get(skill['name'], DEFAULT_QUESTIONS(skill['name']))
        meta = SKILL_META.get(skill['name'], DEFAULT_META)

    data = request.get_json(silent=True) or {}
    reason       = data.get('reason', 'Tab switching limit exceeded')
    time_elapsed = int(data.get('time_elapsed', 0))
    answers      = data.get('answers', {})

    score = sum(
        1 for i, q in enumerate(questions)
        if answers.get(f'q{i}') == q['answer']
    )
    total = len(questions)
    pct   = int((score / total) * 100) if total else 0
    mins, secs = divmod(time_elapsed, 60)

    existing = db.execute(
        'SELECT id FROM user_skills WHERE user_id=? AND skill_id=?',
        [session['user_id'], skill_id]
    ).fetchone()
    if existing:
        db.execute(
            'UPDATE user_skills SET verified=0, score=?, correct_answers=?, '
            'time_taken_secs=?, terminated_reason=? WHERE user_id=? AND skill_id=?',
            [pct, score, time_elapsed, reason, session['user_id'], skill_id]
        )
    else:
        db.execute(
            'INSERT INTO user_skills (user_id, skill_id, verified, score, '
            'correct_answers, time_taken_secs, terminated_reason) VALUES (?,?,0,?,?,?,?)',
            [session['user_id'], skill_id, pct, score, time_elapsed, reason]
        )
    db.commit()

    session['exam_result'] = {
        'skill_id':          skill_id,
        'skill_name':        skill['name'],
        'score':             pct,
        'correct':           score,
        'total':             total,
        'verified':          False,
        'passing':           meta['passing'],
        'duration':          meta['duration'],
        'time_taken':        f'{mins}:{secs:02d}',
        'terminated_reason': reason,
    }
    return jsonify({'ok': True, 'redirect': url_for('exam_result', skill_id=skill_id)})


@app.route('/skills/verify/<int:skill_id>/result')
@candidate_required
def exam_result(skill_id):
    result = session.pop('exam_result', None)
    if not result or result.get('skill_id') != skill_id:
        db = get_db()
        skill = db.execute('SELECT * FROM skills WHERE id=?', [skill_id]).fetchone()
        us = db.execute('SELECT * FROM user_skills WHERE user_id=? AND skill_id=?',
                        [session['user_id'], skill_id]).fetchone()
        if not skill or not us:
            return redirect(url_for('skills_page'))
        meta = dict(SKILL_META.get(skill['name'], DEFAULT_META))
        db_q_count = db.execute(
            'SELECT COUNT(*) FROM skill_questions WHERE skill_id=?', [skill_id]
        ).fetchone()[0]
        if db_q_count:
            meta['questions'] = db_q_count
        tt_secs = us['time_taken_secs']
        time_taken_str = f'{tt_secs // 60}:{tt_secs % 60:02d}' if tt_secs else None
        result = {
            'skill_id': skill_id,
            'skill_name': skill['name'],
            'score': us['score'],
            'correct': us['correct_answers'],
            'total': meta['questions'],
            'verified': bool(us['verified']),
            'passing': meta['passing'],
            'duration': meta['duration'],
            'time_taken': time_taken_str,
        }
    return render_template('exam_result.html', result=result, user=get_current_user())


@app.route('/skills/verify/<int:skill_id>/ended')
@candidate_required
def exam_ended(skill_id):
    data = session.pop('exam_ended_data', None)
    if not data or data.get('skill_id') != skill_id:
        db = get_db()
        skill = db.execute('SELECT * FROM skills WHERE id=?', [skill_id]).fetchone()
        if not skill:
            return redirect(url_for('skills_page'))
        data = {
            'skill_id': skill_id,
            'skill_name': skill['name'],
            'fs_exit_count': 3,
        }
    return render_template('exam_ended.html', data=data, user=get_current_user())


# ── Recruiter: Admin token verification ──────────────────────────────────────

@app.route('/recruiter/verify-admin-token', methods=['POST'])
@admin_required
def verify_admin_token():
    data  = request.get_json(silent=True) or {}
    token = data.get('token', '').strip()

    if not _ADMIN_TOKEN:
        return jsonify({'ok': False, 'error': 'Admin token is not configured on this server.'})

    if token == _ADMIN_TOKEN:
        session['admin_token_verified'] = True
        return jsonify({'ok': True})

    # Wrong token — send alert, log out, respond
    _send_security_alert(
        attempted_email=session.get('email', 'Unknown'),
        ip=request.headers.get('X-Forwarded-For', request.remote_addr),
        user_agent=request.headers.get('User-Agent', ''),
        page='Manage Questions',
    )
    session.clear()
    return jsonify({'ok': False, 'logout': True})


# ── Recruiter: Skill Question Management ─────────────────────────────────────

@app.route('/recruiter/questions')
@admin_required
def manage_questions():
    if not session.get('admin_token_verified'):
        flash('Admin token verification required.', 'error')
        return redirect(url_for('recruiter_dashboard') + '?verify=questions')
    db = get_db()
    skills = db.execute(
        '''SELECT s.id, s.name, s.category,
                  COUNT(sq.id) as question_count
           FROM skills s
           LEFT JOIN skill_questions sq ON s.id = sq.skill_id
           GROUP BY s.id
           ORDER BY s.category, s.name'''
    ).fetchall()
    return render_template('manage_questions.html', skills=skills,
                           skill=None, questions=None,
                           user=get_current_user(), profile=_get_recruiter_profile())


@app.route('/recruiter/questions/<int:skill_id>', methods=['GET', 'POST'])
@admin_required
def manage_skill_questions(skill_id):
    if not session.get('admin_token_verified'):
        flash('Admin token verification required.', 'error')
        return redirect(url_for('recruiter_dashboard') + '?verify=questions')
    db = get_db()
    skill = db.execute('SELECT * FROM skills WHERE id=?', [skill_id]).fetchone()
    if not skill:
        flash('Skill not found.', 'error')
        return redirect(url_for('manage_questions'))

    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            q_text = request.form.get('question_text', '').strip()
            opt_a  = request.form.get('option_a', '').strip()
            opt_b  = request.form.get('option_b', '').strip()
            opt_c  = request.form.get('option_c', '').strip()
            opt_d  = request.form.get('option_d', '').strip()
            correct = request.form.get('correct_option', '').upper()
            marks   = request.form.get('marks', 1, type=int)
            diff    = request.form.get('difficulty', 'Medium')
            if not all([q_text, opt_a, opt_b, opt_c, opt_d]) or correct not in ('A', 'B', 'C', 'D'):
                flash('All fields are required and correct answer must be A, B, C, or D.', 'error')
            elif diff not in ('Easy', 'Medium', 'Hard'):
                flash('Difficulty must be Easy, Medium, or Hard.', 'error')
            else:
                db.execute(
                    '''INSERT INTO skill_questions
                       (skill_id, question_text, option_a, option_b, option_c, option_d,
                        correct_option, marks, difficulty, created_by)
                       VALUES (?,?,?,?,?,?,?,?,?,?)''',
                    [skill_id, q_text, opt_a, opt_b, opt_c, opt_d,
                     correct, marks, diff, session['user_id']]
                )
                db.commit()
                flash('Question added successfully.', 'success')
        elif action == 'delete':
            q_id = request.form.get('question_id', type=int)
            if q_id:
                db.execute('DELETE FROM skill_questions WHERE id=? AND skill_id=?',
                           [q_id, skill_id])
                db.commit()
                flash('Question deleted.', 'success')
        return redirect(url_for('manage_skill_questions', skill_id=skill_id))

    questions = db.execute(
        'SELECT * FROM skill_questions WHERE skill_id=? ORDER BY id', [skill_id]
    ).fetchall()
    skills = db.execute(
        '''SELECT s.id, s.name, s.category,
                  COUNT(sq.id) as question_count
           FROM skills s
           LEFT JOIN skill_questions sq ON s.id = sq.skill_id
           GROUP BY s.id
           ORDER BY s.category, s.name'''
    ).fetchall()
    return render_template('manage_questions.html', skill=skill, questions=questions,
                           skills=skills, user=get_current_user(),
                           profile=_get_recruiter_profile())


_EXCEL_REQUIRED_COLS = [
    'Skill Name', 'Question', 'Option A', 'Option B', 'Option C', 'Option D',
    'Correct Answer', 'Difficulty Level', 'Experience Level',
]
_VALID_DIFFICULTIES = {'easy', 'medium', 'hard'}
_VALID_ANSWERS      = {'a', 'b', 'c', 'd'}
_VALID_EXPERIENCE   = {'fresher', 'junior', 'mid', 'senior', 'all'}


def _parse_excel_rows(file_stream):
    """Return (rows, error_string). rows is list of dicts with keys matching _EXCEL_REQUIRED_COLS."""
    if not _OPENPYXL_OK:
        return None, 'openpyxl is not installed on this server.'
    try:
        wb = openpyxl.load_workbook(file_stream, read_only=True, data_only=True)
        ws = wb.active
        raw_headers = [str(c.value).strip() if c.value is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
    except Exception as exc:
        return None, f'Could not read Excel file: {exc}'

    missing = [h for h in _EXCEL_REQUIRED_COLS if h not in raw_headers]
    if missing:
        return None, f'Missing required columns: {", ".join(missing)}'

    col_idx = {h: raw_headers.index(h) for h in _EXCEL_REQUIRED_COLS}
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or str(v).strip() == '' for v in row):
            continue
        r = {h: (str(row[col_idx[h]]).strip() if row[col_idx[h]] is not None else '') for h in _EXCEL_REQUIRED_COLS}
        r['_row'] = i
        rows.append(r)
    return rows, None


def _validate_excel_row(r, skill_name_to_id, existing_questions):
    errors = []
    skill_id = skill_name_to_id.get(r['Skill Name'].lower())
    if not skill_id:
        errors.append(f'Unknown skill "{r["Skill Name"]}"')
    if not r['Question']:
        errors.append('Question is empty')
    for opt in ('Option A', 'Option B', 'Option C', 'Option D'):
        if not r[opt]:
            errors.append(f'{opt} is empty')
    if r['Correct Answer'].upper() not in ('A', 'B', 'C', 'D'):
        errors.append(f'Correct Answer must be A/B/C/D, got "{r["Correct Answer"]}"')
    if r['Difficulty Level'].lower() not in _VALID_DIFFICULTIES:
        errors.append(f'Difficulty must be Easy/Medium/Hard, got "{r["Difficulty Level"]}"')
    exp = r['Experience Level'].lower() if r['Experience Level'] else 'all'
    if exp not in _VALID_EXPERIENCE:
        errors.append(f'Experience Level must be Fresher/Junior/Mid/Senior/All, got "{r["Experience Level"]}"')

    is_dup = False
    if skill_id and r['Question']:
        key = (skill_id, r['Question'].lower())
        if key in existing_questions:
            is_dup = True

    return skill_id, errors, is_dup


@app.route('/recruiter/questions/upload-excel', methods=['POST'])
@admin_required
@csrf.exempt
def upload_questions_excel():
    if not _OPENPYXL_OK:
        return jsonify({'ok': False, 'error': 'openpyxl not installed on server.'}), 500

    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'ok': False, 'error': 'No file uploaded.'}), 400
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ('.xlsx', '.xls'):
        return jsonify({'ok': False, 'error': 'Only .xlsx and .xls files are accepted.'}), 400

    rows, err = _parse_excel_rows(f.stream)
    if err:
        return jsonify({'ok': False, 'error': err}), 400

    if not rows:
        return jsonify({'ok': False, 'error': 'The file contains no data rows.'}), 400

    db = get_db()
    skill_rows = db.execute('SELECT id, name FROM skills').fetchall()
    skill_name_to_id = {s['name'].lower(): s['id'] for s in skill_rows}

    existing = db.execute(
        'SELECT skill_id, LOWER(question_text) as qt FROM skill_questions'
    ).fetchall()
    existing_questions = {(e['skill_id'], e['qt']) for e in existing}

    preview = []
    for r in rows:
        skill_id, errors, is_dup = _validate_excel_row(r, skill_name_to_id, existing_questions)
        status = 'dup' if (not errors and is_dup) else ('error' if errors else 'ok')
        preview.append({
            'row':        r['_row'],
            'skill':      r['Skill Name'],
            'question':   r['Question'],
            'option_a':   r['Option A'],
            'option_b':   r['Option B'],
            'option_c':   r['Option C'],
            'option_d':   r['Option D'],
            'correct':    r['Correct Answer'].upper() if r['Correct Answer'] else '',
            'difficulty': r['Difficulty Level'],
            'experience': r['Experience Level'] or 'All',
            'status':     status,
            'errors':     errors,
        })

    return jsonify({'ok': True, 'rows': preview})


@app.route('/recruiter/questions/import-excel', methods=['POST'])
@admin_required
@csrf.exempt
def import_questions_excel():
    payload = request.get_json(silent=True) or {}
    rows = payload.get('rows', [])
    if not rows:
        return jsonify({'ok': False, 'error': 'No rows provided.'}), 400

    db = get_db()
    skill_rows = db.execute('SELECT id, name FROM skills').fetchall()
    skill_name_to_id = {s['name'].lower(): s['id'] for s in skill_rows}

    existing = db.execute(
        'SELECT skill_id, LOWER(question_text) as qt FROM skill_questions'
    ).fetchall()
    existing_questions = {(e['skill_id'], e['qt']) for e in existing}

    imported = 0
    skipped  = 0
    for r in rows:
        r_norm = {
            'Skill Name':      r.get('skill', ''),
            'Question':        r.get('question', ''),
            'Option A':        r.get('option_a', ''),
            'Option B':        r.get('option_b', ''),
            'Option C':        r.get('option_c', ''),
            'Option D':        r.get('option_d', ''),
            'Correct Answer':  r.get('correct', ''),
            'Difficulty Level': r.get('difficulty', 'Medium'),
            'Experience Level': r.get('experience', 'All'),
            '_row': 0,
        }
        skill_id, errors, is_dup = _validate_excel_row(r_norm, skill_name_to_id, existing_questions)
        if errors or is_dup or not skill_id:
            skipped += 1
            continue
        exp = r_norm['Experience Level'].strip() or 'All'
        db.execute(
            '''INSERT INTO skill_questions
               (skill_id, question_text, option_a, option_b, option_c, option_d,
                correct_option, marks, difficulty, experience_level, created_by)
               VALUES (?,?,?,?,?,?,?,1,?,?,?)''',
            [skill_id, r_norm['Question'], r_norm['Option A'], r_norm['Option B'],
             r_norm['Option C'], r_norm['Option D'], r_norm['Correct Answer'].upper(),
             r_norm['Difficulty Level'].capitalize(), exp, session['user_id']]
        )
        existing_questions.add((skill_id, r_norm['Question'].lower()))
        imported += 1
    db.commit()
    return jsonify({'ok': True, 'imported': imported, 'skipped': skipped})


@app.route('/recruiter/questions/download-template')
@admin_required
def download_questions_template():
    if not _OPENPYXL_OK:
        flash('openpyxl is not installed on this server.', 'error')
        return redirect(url_for('manage_questions'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Questions'
    headers = _EXCEL_REQUIRED_COLS[:]
    ws.append(headers)
    ws.append([
        'Python', 'What is the output of print(type([]))?',
        "<class 'list'>", "<class 'tuple'>", "<class 'dict'>", "<class 'set'>",
        'A', 'Easy', 'All',
    ])
    ws.append([
        'JavaScript', 'Which keyword declares a block-scoped variable?',
        'var', 'let', 'const', 'function',
        'B', 'Medium', 'Junior',
    ])

    from openpyxl.styles import Font, PatternFill
    hdr_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    hdr_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        ws.column_dimensions[cell.column_letter].width = max(18, len(str(cell.value)) + 4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='skill_questions_template.xlsx',
    )


def _get_recruiter_profile():
    db = get_db()
    return db.execute(
        'SELECT * FROM recruiter_profiles WHERE user_id=?', [session['user_id']]
    ).fetchone()


# ── Registration type selector ────────────────────────────────────────────────

@app.route('/register')
def register():
    if session.get('user_id'):
        return redirect(url_for('candidate_profile' if session.get('role') == 'candidate' else 'recruiter_dashboard'))
    return render_template('register.html', user=None)


# ── Candidate auth ────────────────────────────────────────────────────────────

@app.route('/candidate/signup', methods=['GET', 'POST'])
@limiter.limit('10 per minute')
def candidate_signup():
    if session.get('role') == 'candidate':
        return redirect(url_for('candidate_profile'))

    all_skills = get_db().execute('SELECT * FROM skills ORDER BY category, name').fetchall()

    if request.method == 'POST':
        name        = request.form.get('name', '').strip()
        email       = request.form.get('email', '').strip().lower()
        _phone_code = request.form.get('phone_code', '').strip()
        _phone_num  = request.form.get('phone', '').strip()
        phone       = ((_phone_code + ' ' + _phone_num).strip()) if _phone_num else ''
        job_title   = request.form.get('job_title', '').strip()
        experience  = request.form.get('experience', '').strip()
        headline    = request.form.get('headline', '').strip()
        location    = request.form.get('location', '').strip()
        bio         = request.form.get('bio', '').strip()
        linkedin    = request.form.get('linkedin', '').strip()
        github      = request.form.get('github', '').strip()
        skill_ids   = request.form.getlist('skills', type=int)
        work_status = request.form.get('work_status', '').strip()
        password    = request.form.get('password', '')
        confirm     = request.form.get('confirm_password', '')
        terms       = request.form.get('terms')

        errors = []
        if not name:        errors.append('Full name is required.')
        if not email:       errors.append('Email address is required.')
        if not work_status: errors.append('Please select your work status.')
        if not terms:       errors.append('You must accept the terms.')
        pw_err = validate_password(password)
        if pw_err:     errors.append(pw_err)
        if not pw_err and password != confirm:
            errors.append('Passwords do not match.')
        if not errors and get_db().execute('SELECT id FROM users WHERE email=?', [email]).fetchone():
            errors.append('An account with this email already exists.')

        if errors:
            for e in errors:
                flash(e, 'error')
            return render_template('candidate_signup.html', user=None, all_skills=all_skills,
                                   form=request.form, sel_skills=skill_ids)

        # Handle resume upload
        resume_filename = None
        resume_file = request.files.get('resume')
        if resume_file and resume_file.filename:
            if not allowed_resume(resume_file.filename):
                flash('Resume must be PDF, DOC, or DOCX.', 'error')
                return render_template('candidate_signup.html', user=None, all_skills=all_skills,
                                       form=request.form, sel_skills=skill_ids)
            if len(resume_file.read()) > MAX_RESUME_SIZE:
                flash('Only PDF, DOC, and DOCX resumes up to 1 MB are allowed.', 'error')
                return render_template('candidate_signup.html', user=None, all_skills=all_skills,
                                       form=request.form, sel_skills=skill_ids)
            resume_file.seek(0)
            if not validate_file_magic(resume_file, ALLOWED_RESUME_EXT):
                flash('Resume file content does not match the declared format.', 'error')
                return render_template('candidate_signup.html', user=None, all_skills=all_skills,
                                       form=request.form, sel_skills=skill_ids)
            os.makedirs(RESUME_UPLOAD_FOLDER, exist_ok=True)
            resume_filename = secure_filename(f'{email}_{resume_file.filename}')
            resume_file.save(os.path.join(RESUME_UPLOAD_FOLDER, resume_filename))

        db = get_db()
        raw_token, token_hash, expires_at = _gen_email_token()
        pw_hash = generate_password_hash(password)
        try:
            cur = db.execute(
                'INSERT INTO users (name, email, password_hash, role, email_verified,'
                ' email_verification_token_hash, email_verification_expires_at, email_verification_sent_at)'
                ' VALUES (?,?,?,?,0,?,?,?)',
                [name, email, pw_hash, 'candidate', token_hash, expires_at, datetime.utcnow()]
            )
        except Exception:
            # Fallback: schema migration hasn't run yet — insert without verification columns
            raw_token = None
            cur = db.execute(
                'INSERT INTO users (name, email, password_hash, role) VALUES (?,?,?,?)',
                [name, email, pw_hash, 'candidate']
            )
        uid = cur.lastrowid
        db.execute('''INSERT INTO candidate_profiles
                      (user_id, headline, location, bio, linkedin, github, phone, job_title, experience, resume_filename, work_status)
                      VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
                   [uid, headline, location, bio, linkedin, github, phone, job_title, experience, resume_filename, work_status])
        for sid in skill_ids:
            db.execute('INSERT OR IGNORE INTO user_skills (user_id, skill_id) VALUES (?,?)', [uid, sid])
        db.commit()

        if raw_token:
            send_verification_email(email, name, raw_token)
        session.update({'user_id': uid, 'role': 'candidate', 'name': name,
                        'email_verified': not raw_token, 'pending_email': email, 'pending_name': name})
        return redirect(url_for('email_sent') if raw_token else url_for('jobs'))

    return render_template('candidate_signup.html', user=None, all_skills=all_skills,
                           form={}, sel_skills=[])


# ── Social login helper ──────────────────────────────────────────────────────

def _oauth_state(role):
    """Return a state string that embeds the role for recovery after redirect."""
    return f'{role}:{secrets.token_urlsafe(16)}'


def _role_from_state(state_str, fallback='candidate'):
    """Extract role from a state string like 'candidate:abc123'."""
    try:
        part = (state_str or '').split(':', 1)[0]
        return part if part in ('candidate', 'recruiter') else fallback
    except Exception:
        return fallback


def _handle_social_login(email, name, google_id, linkedin_id, role):
    """Find or create user from social OAuth, set session, redirect."""
    provider = 'google' if google_id else 'linkedin'
    app.logger.info('[OAUTH] provider=%s email=%r role=%r name=%r', provider, email, role, name)

    db = get_db()
    login_url = url_for('candidate_login') if role == 'candidate' else url_for('recruiter_login')

    if not email:
        app.logger.error('[OAUTH_ERROR] provider=%s no email returned from provider', provider)
        flash('Social login failed: could not retrieve your email. Please try again.', 'error')
        return redirect(login_url)

    user = db.execute('SELECT * FROM users WHERE email=?', [email]).fetchone()
    app.logger.info('[OAUTH] existing_user=%s', bool(user))

    if user:
        # Strict role enforcement: block cross-role OAuth login
        if user['role'] != role:
            app.logger.warning(
                '[AUTH_ROLE_CHECK] email=%s stored_role=%s requested_area=oauth_%s_login allowed=false reason=role_mismatch',
                email, user['role'], role)
            qs = urlencode({'oauth_mismatch': user['role'], 'oauth_email': email})
            return redirect(f'{login_url}?{qs}')

        # Update social ID if not already stored
        if google_id and not user['google_id']:
            try:
                db.execute('UPDATE users SET google_id=? WHERE id=?', [google_id, user['id']])
                db.commit()
                app.logger.info('[OAUTH] google_id linked for user_id=%s', user['id'])
            except Exception as exc:
                app.logger.warning('[OAUTH] could not link google_id: %s', exc)
        if linkedin_id and not user['linkedin_id']:
            try:
                db.execute('UPDATE users SET linkedin_id=? WHERE id=?', [linkedin_id, user['id']])
                db.commit()
                app.logger.info('[OAUTH] linkedin_id linked for user_id=%s', user['id'])
            except Exception as exc:
                app.logger.warning('[OAUTH] could not link linkedin_id: %s', exc)

    else:
        # New user: only candidates may be auto-created via OAuth.
        # Recruiter accounts must be created through the registration form.
        if role == 'recruiter':
            app.logger.warning(
                '[AUTH_ROLE_CHECK] email=%s stored_role=none requested_area=oauth_recruiter_login '
                'allowed=false reason=no_recruiter_oauth_signup', email)
            flash('No recruiter account found for this email. '
                  'Please register as a recruiter or log in with your email and password.', 'error')
            return redirect(url_for('recruiter_login'))

        # Auto-create candidate account
        placeholder_hash = generate_password_hash(secrets.token_hex(32))
        app.logger.info('[OAUTH] creating new candidate user email=%r', email)
        try:
            db.execute(
                'INSERT INTO users (name, email, password_hash, role, email_verified, google_id, linkedin_id)'
                ' VALUES (?,?,?,?,1,?,?)',
                [name or email.split('@')[0], email, placeholder_hash, 'candidate', google_id, linkedin_id]
            )
            db.commit()
        except Exception as exc:
            app.logger.error('[OAUTH_ERROR] INSERT users failed email=%r: %s', email, exc)
            flash('Social login failed. Please try again.', 'error')
            return redirect(login_url)

        user = db.execute('SELECT * FROM users WHERE email=?', [email]).fetchone()
        if not user:
            app.logger.error('[OAUTH_ERROR] user not found after INSERT email=%r', email)
            flash('Social login failed. Please try again.', 'error')
            return redirect(login_url)

        app.logger.info('[OAUTH] user_created=True user_id=%s role=candidate', user['id'])

        try:
            db.execute('INSERT INTO candidate_profiles (user_id) VALUES (?)', [user['id']])
            db.commit()
            app.logger.info('[OAUTH] candidate_profile created for user_id=%s', user['id'])
        except Exception as exc:
            app.logger.warning('[OAUTH] profile insert skipped (may already exist): %s', exc)

    # Final DB role verification — belt-and-suspenders before session write
    db_role = user['role']
    if db_role != role:
        app.logger.error(
            '[AUTH_ROLE_CHECK] email=%s stored_role=%s requested_area=oauth_%s_login '
            'allowed=false reason=db_role_mismatch_at_session_write', email, db_role, role)
        flash('Access denied.', 'error')
        return redirect(login_url)

    app.logger.info('[AUTH_ROLE_CHECK] email=%s stored_role=%s requested_area=oauth_%s_login allowed=true',
                    email, db_role, role)

    # Set session
    app.logger.info('[OAUTH] setting session for user_id=%s role=%s', user['id'], role)
    if role == 'candidate':
        try:
            cp = db.execute(
                'SELECT profile_photo, headline FROM candidate_profiles WHERE user_id=?',
                [user['id']]
            ).fetchone()
        except Exception as exc:
            app.logger.warning('[OAUTH] candidate_profiles fetch failed: %s', exc)
            cp = None
        session.clear()
        session.update({
            'user_id':       user['id'],
            'role':          'candidate',
            'name':          user['name'],
            'email':         user['email'],
            'profile_photo': cp['profile_photo'] if cp else None,
            'headline':      (cp.get('headline') or '') if cp else '',
            'email_verified': True,
        })
        app.logger.info('[OAUTH] login_success=True redirect=/candidate/dashboard')
        flash(f'Welcome, {user["name"]}!', 'success')
        return redirect(url_for('candidate_dashboard'))
    else:
        try:
            rp = db.execute(
                'SELECT profile_photo, company FROM recruiter_profiles WHERE user_id=?',
                [user['id']]
            ).fetchone()
        except Exception as exc:
            app.logger.warning('[OAUTH] recruiter_profiles fetch failed: %s', exc)
            rp = None
        session.clear()
        session.update({
            'user_id':       user['id'],
            'role':          'recruiter',
            'name':          user['name'],
            'email':         user['email'],
            'profile_photo': rp['profile_photo'] if rp else None,
            'headline':      (rp.get('company') or '') if rp else '',
            'email_verified': True,
        })
        app.logger.info('[OAUTH] login_success=True redirect=/recruiter/dashboard')
        flash(f'Welcome, {user["name"]}!', 'success')
        return redirect(url_for('recruiter_dashboard'))


# ── Resend OTP ───────────────────────────────────────────────────────────────

@app.route('/candidate/resend-otp', methods=['POST'])
@limiter.limit('5 per hour')
@csrf.exempt
def candidate_resend_otp():
    data = request.get_json(silent=True) or {}
    email = data.get('email', '').strip().lower()
    if not email:
        return jsonify({'ok': False, 'error': 'Please enter your email address first.'})
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE email=? AND role='candidate'", [email]).fetchone()
    if not user:
        return jsonify({'ok': True, 'message': 'If this email is registered, a verification link has been sent.'})
    if user.get('email_verified', 0):
        return jsonify({'ok': True, 'message': 'Your email is already verified. You can log in now.'})
    try:
        sent_at = user['email_verification_sent_at']
        if sent_at:
            if isinstance(sent_at, str):
                sent_at = datetime.strptime(sent_at[:19], '%Y-%m-%d %H:%M:%S')
            seconds_ago = (datetime.utcnow() - sent_at).total_seconds()
            if seconds_ago < 60:
                wait = int(60 - seconds_ago)
                return jsonify({'ok': False, 'error': f'Please wait {wait}s before requesting another OTP.', 'wait': wait})
    except Exception:
        pass
    raw_token, token_hash, expires_at = _gen_email_token()
    db.execute(
        'UPDATE users SET email_verification_token_hash=?, email_verification_expires_at=?,'
        ' email_verification_sent_at=? WHERE id=?',
        [token_hash, expires_at, datetime.utcnow(), user['id']]
    )
    db.commit()
    try:
        send_verification_email(user['email'], user['name'], raw_token)
        return jsonify({'ok': True, 'message': 'OTP has been resent to your registered email.'})
    except Exception:
        return jsonify({'ok': False, 'error': 'Unable to resend OTP. Please try again.'})


@app.route('/recruiter/resend-otp', methods=['POST'])
@limiter.limit('5 per hour')
@csrf.exempt
def recruiter_resend_otp():
    data = request.get_json(silent=True) or {}
    email = data.get('email', '').strip().lower()
    if not email:
        return jsonify({'ok': False, 'error': 'Please enter your email address first.'})
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE email=? AND role='recruiter'", [email]).fetchone()
    if not user:
        return jsonify({'ok': True, 'message': 'If this email is registered, a verification link has been sent.'})
    if user.get('email_verified', 0):
        return jsonify({'ok': True, 'message': 'Your email is already verified. You can log in now.'})
    try:
        sent_at = user['email_verification_sent_at']
        if sent_at:
            if isinstance(sent_at, str):
                sent_at = datetime.strptime(sent_at[:19], '%Y-%m-%d %H:%M:%S')
            seconds_ago = (datetime.utcnow() - sent_at).total_seconds()
            if seconds_ago < 60:
                wait = int(60 - seconds_ago)
                return jsonify({'ok': False, 'error': f'Please wait {wait}s before requesting another OTP.', 'wait': wait})
    except Exception:
        pass
    raw_token, token_hash, expires_at = _gen_email_token()
    db.execute(
        'UPDATE users SET email_verification_token_hash=?, email_verification_expires_at=?,'
        ' email_verification_sent_at=? WHERE id=?',
        [token_hash, expires_at, datetime.utcnow(), user['id']]
    )
    db.commit()
    try:
        send_verification_email(user['email'], user['name'], raw_token)
        return jsonify({'ok': True, 'message': 'OTP has been resent to your registered email.'})
    except Exception:
        return jsonify({'ok': False, 'error': 'Unable to resend OTP. Please try again.'})


# ── Google OAuth ─────────────────────────────────────────────────────────────

@app.route('/candidate/google-login')
def candidate_google_login():
    if not _GOOGLE_CLIENT_ID:
        return redirect(url_for('candidate_login') + '?social_error=google')
    state = _oauth_state('candidate')
    session['oauth_role'] = 'candidate'
    app.logger.info('[OAUTH] initiating google login role=candidate')
    return _oauth.google.authorize_redirect(url_for('google_callback', _external=True), state=state)


@app.route('/recruiter/google-login')
def recruiter_google_login():
    app.logger.warning('[AUTH_ROLE_CHECK] recruiter Google OAuth blocked — social login disabled for recruiters')
    return redirect(url_for('recruiter_login'))


@app.route('/auth/google/callback')
def google_callback():
    # Extract role from state param first (doesn't depend on session)
    raw_state = request.args.get('state', '')
    role = _role_from_state(raw_state, fallback=session.pop('oauth_role', 'candidate'))
    app.logger.info('[OAUTH] google_callback state=%r role=%s', raw_state, role)
    # Social login is disabled for recruiters — block at callback level too
    if role == 'recruiter':
        app.logger.warning('[AUTH_ROLE_CHECK] recruiter role in Google callback blocked email=unknown')
        flash('Social login is not available for recruiters. Please use email and password.', 'error')
        return redirect(url_for('recruiter_login'))
    try:
        token     = _oauth.google.authorize_access_token()
        user_info = token.get('userinfo') or {}
        email     = user_info.get('email', '').strip().lower()
        name      = user_info.get('name') or email.split('@')[0]
        google_id = user_info.get('sub', '')
        app.logger.info('[OAUTH] google userinfo email=%r google_id=%r name=%r', email, google_id, name)
    except Exception as exc:
        app.logger.error('[OAUTH_ERROR] google authorize_access_token failed: %s', exc)
        flash('Google login failed. Please try again.', 'error')
        return redirect(url_for('candidate_login' if role == 'candidate' else 'recruiter_login'))
    return _handle_social_login(email, name, google_id, None, role)


# ── LinkedIn OAuth ───────────────────────────────────────────────────────────

def _linkedin_auth_url(role):
    """Build LinkedIn authorization redirect URL with role encoded in state."""
    state = _oauth_state(role)
    session['_li_state'] = state
    session['oauth_role'] = role
    params = urlencode({
        'response_type': 'code',
        'client_id':     _LINKEDIN_CLIENT_ID,
        'redirect_uri':  url_for('linkedin_callback', _external=True),
        'state':         state,
        'scope':         'openid profile email',
    })
    return f'https://www.linkedin.com/oauth/v2/authorization?{params}'


@app.route('/candidate/linkedin-login')
def candidate_linkedin_login():
    if not _LINKEDIN_CLIENT_ID:
        return redirect(url_for('candidate_login') + '?social_error=linkedin')
    app.logger.info('[OAUTH] initiating linkedin login role=candidate')
    return redirect(_linkedin_auth_url('candidate'))


@app.route('/recruiter/linkedin-login')
def recruiter_linkedin_login():
    app.logger.warning('[AUTH_ROLE_CHECK] recruiter LinkedIn OAuth blocked — social login disabled for recruiters')
    return redirect(url_for('recruiter_login'))


@app.route('/auth/linkedin/callback')
def linkedin_callback():
    import requests as _req

    raw_state    = request.args.get('state', '')
    role         = _role_from_state(raw_state, fallback=session.pop('oauth_role', 'candidate'))
    expected     = session.pop('_li_state', None)
    login_url    = url_for('candidate_login' if role == 'candidate' else 'recruiter_login')

    app.logger.info('[OAUTH] linkedin_callback state=%r role=%s', raw_state, role)
    # Social login is disabled for recruiters — block at callback level too
    if role == 'recruiter':
        app.logger.warning('[AUTH_ROLE_CHECK] recruiter role in LinkedIn callback blocked email=unknown')
        flash('Social login is not available for recruiters. Please use email and password.', 'error')
        return redirect(url_for('recruiter_login'))

    # CSRF state check
    if not raw_state or raw_state != expected:
        app.logger.error('[OAUTH_ERROR] linkedin state mismatch expected=%r got=%r', expected, raw_state)
        flash('LinkedIn login failed: security check failed. Please try again.', 'error')
        return redirect(login_url)

    code = request.args.get('code')
    if not code:
        app.logger.error('[OAUTH_ERROR] linkedin no code in callback error=%r', request.args.get('error'))
        flash('LinkedIn login failed. Please try again.', 'error')
        return redirect(login_url)

    try:
        # Step 1: exchange code for access token
        tok_resp = _req.post(
            'https://www.linkedin.com/oauth/v2/accessToken',
            data={
                'grant_type':   'authorization_code',
                'code':         code,
                'redirect_uri': url_for('linkedin_callback', _external=True),
                'client_id':    _LINKEDIN_CLIENT_ID,
                'client_secret': _LINKEDIN_CLIENT_SECRET,
            },
            timeout=15,
        )
        tok_data     = tok_resp.json()
        access_token = tok_data.get('access_token', '')
        app.logger.info('[OAUTH] linkedin token exchange HTTP=%s has_token=%s',
                        tok_resp.status_code, bool(access_token))
        if not access_token:
            raise ValueError(f'No access_token in response: {tok_data}')

        # Step 2: fetch userinfo from LinkedIn OIDC endpoint
        ui_resp   = _req.get(
            'https://api.linkedin.com/v2/userinfo',
            headers={'Authorization': f'Bearer {access_token}'},
            timeout=10,
        )
        user_info   = ui_resp.json()
        app.logger.info('[OAUTH] linkedin userinfo HTTP=%s keys=%s',
                        ui_resp.status_code, list(user_info.keys()))

        email       = user_info.get('email', '').strip().lower()
        name        = user_info.get('name') or email.split('@')[0]
        linkedin_id = user_info.get('sub', '')
        app.logger.info('[OAUTH] linkedin resolved email=%r linkedin_id=%r name=%r',
                        email, linkedin_id, name)

    except Exception as exc:
        app.logger.error('[OAUTH_ERROR] linkedin flow failed: %s', exc, exc_info=True)
        flash('LinkedIn login failed. Please try again.', 'error')
        return redirect(login_url)

    return _handle_social_login(email, name, None, linkedin_id, role)


@app.route('/candidate/login', methods=['GET', 'POST'])
@limiter.limit('15 per minute')
def candidate_login():
    if session.get('role') == 'candidate':
        return redirect(url_for('jobs'))

    email = ''
    mismatch = None
    if request.method == 'POST':
        is_ajax = request.is_json
        if is_ajax:
            _d    = request.get_json(silent=True) or {}
            email = _d.get('email', '').strip().lower()
            password = _d.get('password', '')
        else:
            email    = request.form.get('email', '').strip().lower()
            password = request.form.get('password', '')

        _any = get_db().execute("SELECT role FROM users WHERE email=?", [email]).fetchone()
        actual_role = _any['role'] if _any else None
        app.logger.info('[LOGIN:candidate] email=%r login_type=candidate actual_role=%r', email, actual_role)

        if _any and actual_role != 'candidate':
            app.logger.info('[LOGIN:candidate] role_mismatch=True actual_role=%r', actual_role)
            if is_ajax:
                return jsonify({'success': False, 'role_mismatch': True, 'actual_role': actual_role})
            mismatch = actual_role
            return render_template('candidate_login.html', user=None, email=email, mismatch=mismatch)

        user = get_db().execute(
            "SELECT * FROM users WHERE email=? AND role='candidate'", [email]
        ).fetchone()
        app.logger.info('[LOGIN:candidate] user_found=%s', user is not None)
        pw_ok = bool(user) and check_password_hash(user['password_hash'], password)
        app.logger.info('[LOGIN:candidate] password_check=%s', pw_ok)
        if pw_ok:
            if not user.get('email_verified', 0):
                app.logger.info('[LOGIN:candidate] email_verified=false — blocking login')
                if is_ajax:
                    return jsonify({'success': False, 'verification_required': True,
                                    'email_verified': False,
                                    'error': 'Please verify your email before logging in.'})
                flash('Please verify your email before logging in. Check your inbox or resend the verification link.', 'error')
                return render_template('candidate_login.html', user=None, email=email,
                                       mismatch=None, verification_required=True)
            try:
                cp = get_db().execute(
                    'SELECT profile_photo, headline FROM candidate_profiles WHERE user_id=?',
                    [user['id']]).fetchone()
            except Exception:
                cp = get_db().execute(
                    'SELECT * FROM candidate_profiles WHERE user_id=?',
                    [user['id']]).fetchone()
            session.update({'user_id': user['id'], 'role': 'candidate', 'name': user['name'],
                            'email': user['email'],
                            'profile_photo': cp['profile_photo'] if cp else None,
                            'headline': (cp['headline'] or '') if cp else '',
                            'email_verified': bool(user.get('email_verified', 0))})
            if is_ajax:
                return jsonify({'success': True, 'redirect': url_for('jobs')})
            flash(f'Welcome back, {user["name"]}!', 'success')
            return redirect(url_for('jobs'))

        if is_ajax:
            return jsonify({'success': False, 'error': 'Invalid email or password.'})
        flash('Invalid email or password.', 'error')

    if not email:
        email = request.args.get('email', request.args.get('oauth_email', ''))
    if mismatch is None:
        mismatch = request.args.get('oauth_mismatch')
    return render_template('candidate_login.html', user=None, email=email, mismatch=mismatch)


# ── Candidate notifications ──────────────────────────────────────────────────

@app.route('/candidate/notifications')
@candidate_required
def candidate_notifications():
    db = get_db()
    notifications = db.execute('''
        SELECT n.*, j.title AS job_title,
               rp.company AS company_name
        FROM notifications n
        LEFT JOIN jobs j ON n.job_id = j.id
        LEFT JOIN recruiter_profiles rp ON j.recruiter_id = rp.user_id
        WHERE n.candidate_id=?
        ORDER BY n.created_at DESC
    ''', [session['user_id']]).fetchall()
    return render_template('notifications.html',
                           notifications=notifications,
                           user=get_current_user())


@app.route('/candidate/notifications/<int:notif_id>/read', methods=['POST'])
@candidate_required
def mark_notif_read(notif_id):
    db = get_db()
    db.execute('UPDATE notifications SET is_read=1 WHERE id=? AND candidate_id=?',
               [notif_id, session['user_id']])
    db.commit()
    return redirect(url_for('candidate_notifications'))


@app.route('/candidate/notifications/read-all', methods=['POST'])
@candidate_required
def mark_all_notif_read():
    db = get_db()
    db.execute('UPDATE notifications SET is_read=1 WHERE candidate_id=?',
               [session['user_id']])
    db.commit()
    return redirect(url_for('candidate_notifications'))


# ── Notification JSON API (for panel) ────────────────────────────────────────

@app.route('/api/notifications')
@candidate_required
def api_notifications():
    db = get_db()
    rows = db.execute('''
        SELECT n.id, n.title, n.message, n.type, n.is_read,
               n.redirect_url, n.created_at,
               j.title AS job_title, rp.company AS company_name
        FROM notifications n
        LEFT JOIN jobs j ON n.job_id = j.id
        LEFT JOIN recruiter_profiles rp ON j.recruiter_id = rp.user_id
        WHERE n.candidate_id=?
        ORDER BY n.created_at DESC
        LIMIT 30
    ''', [session['user_id']]).fetchall()
    unread_count = sum(1 for r in rows if not r['is_read'])
    notifs = []
    for r in rows:
        notifs.append({
            'id': r['id'],
            'title': r['title'],
            'message': r['message'],
            'type': r['type'],
            'is_read': bool(r['is_read']),
            'redirect_url': r['redirect_url'] or '',
            'created_at': str(r['created_at']),
            'job_title': r['job_title'] or '',
            'company_name': r['company_name'] or '',
        })
    return jsonify({'notifications': notifs, 'unread_count': unread_count})


@app.route('/api/notifications/<int:notif_id>/read', methods=['POST'])
@candidate_required
def api_notif_read(notif_id):
    db = get_db()
    db.execute('UPDATE notifications SET is_read=1 WHERE id=? AND candidate_id=?',
               [notif_id, session['user_id']])
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/notifications/<int:notif_id>/delete', methods=['POST'])
@candidate_required
def api_notif_delete(notif_id):
    db = get_db()
    db.execute('DELETE FROM notifications WHERE id=? AND candidate_id=?',
               [notif_id, session['user_id']])
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/notifications/read-all', methods=['POST'])
@candidate_required
def api_notif_read_all():
    db = get_db()
    db.execute('UPDATE notifications SET is_read=1 WHERE candidate_id=?',
               [session['user_id']])
    db.commit()
    return jsonify({'ok': True})


# ── Candidate dashboard ───────────────────────────────────────────────────────

@app.route('/candidate/dashboard')
@candidate_required
def candidate_dashboard():
    db = get_db()
    profile = db.execute('SELECT * FROM candidate_profiles WHERE user_id=?',
                         [session['user_id']]).fetchone()
    my_skills = db.execute('''
        SELECT s.*, us.verified, us.score, us.added_at
        FROM user_skills us JOIN skills s ON us.skill_id = s.id
        WHERE us.user_id=? ORDER BY us.verified DESC, s.name
    ''', [session['user_id']]).fetchall()

    applications = db.execute('''
        SELECT a.*, j.title, j.location, j.job_type, j.recruiter_id,
               rp.company AS company_name
        FROM applications a JOIN jobs j ON a.job_id = j.id
        JOIN recruiter_profiles rp ON j.recruiter_id = rp.user_id
        WHERE a.candidate_id=? ORDER BY a.applied_at DESC
    ''', [session['user_id']]).fetchall()

    recommended = db.execute('''
        SELECT j.*, rp.company AS company_name,
               (SELECT COUNT(*) FROM job_skills js
                JOIN user_skills us ON js.skill_id=us.skill_id
                WHERE js.job_id=j.id AND us.user_id=?) AS skill_match
        FROM jobs j JOIN recruiter_profiles rp ON j.recruiter_id=rp.user_id
        WHERE j.active=1
          AND j.id NOT IN (SELECT job_id FROM applications WHERE candidate_id=?)
        ORDER BY skill_match DESC, j.created_at DESC LIMIT 5
    ''', [session['user_id'], session['user_id']]).fetchall()

    return render_template('candidate_dashboard.html',
                           user=get_current_user(), profile=profile,
                           my_skills=my_skills, applications=applications,
                           recommended=recommended)


@app.route('/candidate/profile', methods=['GET', 'POST'])
@candidate_required
def candidate_profile():
    db = get_db()
    if not get_current_user():
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('candidate_login'))

    profile = db.execute('SELECT * FROM candidate_profiles WHERE user_id=?',
                         [session['user_id']]).fetchone()

    if request.method == 'POST':
        # Handle resume delete
        if request.form.get('delete_resume') == '1':
            old = db.execute('SELECT resume_filename FROM candidate_profiles WHERE user_id=?',
                             [session['user_id']]).fetchone()
            if old and old.get('resume_filename'):
                try:
                    os.remove(os.path.join(RESUME_UPLOAD_FOLDER, old['resume_filename']))
                except OSError:
                    pass
            db.execute(
                'UPDATE candidate_profiles SET resume_filename=NULL, resume_data=NULL, resume_original_name=NULL WHERE user_id=?',
                [session['user_id']])
            db.commit()
            flash('Resume deleted.', 'success')
            return redirect(url_for('candidate_profile'))

        name              = request.form.get('name', '').strip()
        headline          = request.form.get('headline', '').strip()
        location          = request.form.get('location', '').strip()
        bio               = request.form.get('bio', '').strip()
        linkedin          = request.form.get('linkedin', '').strip()
        github            = request.form.get('github', '').strip()
        phone             = request.form.get('phone', '').strip()
        job_title         = request.form.get('job_title', '').strip()
        experience        = request.form.get('experience', '').strip()
        gender            = request.form.get('gender', '').strip()
        differently_abled = request.form.get('differently_abled', 'no').strip()

        if not name:
            flash('Name is required.', 'error')
            return redirect(url_for('candidate_profile'))

        db.execute('UPDATE users SET name=? WHERE id=?', [name, session['user_id']])
        db.execute('''UPDATE candidate_profiles
                      SET headline=?, location=?, bio=?, linkedin=?, github=?, phone=?,
                          job_title=?, experience=?, gender=?, differently_abled=?
                      WHERE user_id=?''',
                   [headline, location, bio, linkedin, github, phone,
                    job_title, experience, gender, differently_abled, session['user_id']])
        db.commit()
        session['name'] = name
        flash('Profile updated successfully!', 'success')
        return redirect(url_for('candidate_profile') + '#section-contact')

    my_skills = db.execute('''
        SELECT s.*, us.verified, us.score, us.added_at
        FROM user_skills us JOIN skills s ON us.skill_id = s.id
        WHERE us.user_id=? ORDER BY us.verified DESC, s.name
    ''', [session['user_id']]).fetchall()

    experiences = db.execute(
        'SELECT * FROM candidate_work_experience WHERE user_id=? ORDER BY start_date DESC, id DESC',
        [session['user_id']]).fetchall()
    educations = db.execute(
        'SELECT * FROM candidate_education WHERE user_id=? ORDER BY COALESCE(end_year,\'9999\') DESC, COALESCE(start_year,\'0\') DESC',
        [session['user_id']]).fetchall()
    certifications = db.execute(
        'SELECT * FROM candidate_certifications WHERE user_id=? ORDER BY COALESCE(year,\'0\') DESC, id DESC',
        [session['user_id']]).fetchall()
    try:
        projects = db.execute(
            'SELECT * FROM candidate_projects WHERE user_id=? ORDER BY COALESCE(year,\'0\') DESC, id DESC',
            [session['user_id']]).fetchall()
    except Exception:
        projects = db.execute(
            'SELECT * FROM candidate_projects WHERE user_id=? ORDER BY id DESC',
            [session['user_id']]).fetchall()

    # Calculate total work experience
    now = datetime.now()
    total_months = 0
    for exp in experiences:
        if not exp['start_date']:
            continue
        try:
            sy, sm = int(exp['start_date'][:4]), int(exp['start_date'][5:7])
        except (ValueError, IndexError):
            continue
        if exp['is_current']:
            ey, em = now.year, now.month
        elif exp['end_date']:
            try:
                ey, em = int(exp['end_date'][:4]), int(exp['end_date'][5:7])
            except (ValueError, IndexError):
                continue
        else:
            continue
        diff = (ey - sy) * 12 + (em - sm)
        if diff > 0:
            total_months += diff
    total_exp_years  = total_months // 12
    total_exp_months = total_months % 12

    try:
        applications = db.execute('''
            SELECT a.id, a.status, a.applied_at, a.updated_at,
                   j.title, j.job_type, j.recruiter_id,
                   rp.company AS company_name
            FROM applications a
            JOIN jobs j ON a.job_id = j.id
            JOIN recruiter_profiles rp ON j.recruiter_id = rp.user_id
            WHERE a.candidate_id=? ORDER BY a.applied_at DESC
        ''', [session['user_id']]).fetchall()
    except Exception:
        applications = db.execute('''
            SELECT a.id, a.status, a.applied_at,
                   j.title, j.job_type, j.recruiter_id,
                   rp.company AS company_name
            FROM applications a
            JOIN jobs j ON a.job_id = j.id
            JOIN recruiter_profiles rp ON j.recruiter_id = rp.user_id
            WHERE a.candidate_id=? ORDER BY a.applied_at DESC
        ''', [session['user_id']]).fetchall()

    return render_template('candidate_profile.html',
                           user=get_current_user(), profile=profile, my_skills=my_skills,
                           experiences=experiences, educations=educations,
                           certifications=certifications, projects=projects,
                           total_exp_years=total_exp_years, total_exp_months=total_exp_months,
                           applications=applications)


@app.route('/candidate/profile/upload-resume', methods=['POST'])
@candidate_required
def upload_resume():
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    _err_msg = 'Only PDF, DOC, and DOCX resumes up to 1 MB are allowed.'
    def _err(msg):
        if is_ajax: return jsonify({'ok': False, 'error': msg})
        flash(msg, 'error'); return redirect(url_for('candidate_profile'))
    if 'resume' not in request.files:
        return _err('No file selected.')
    file = request.files['resume']
    if not file or not file.filename:
        return _err('No file selected.')
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in ALLOWED_RESUME_EXT:
        return _err(_err_msg)
    file.seek(0, 2); size = file.tell(); file.seek(0)
    if size > MAX_RESUME_SIZE:
        return _err(_err_msg)
    if not validate_file_magic(file, ALLOWED_RESUME_EXT):
        return _err('File content does not match the declared format.')
    file_bytes = file.read()
    encoded = base64.b64encode(file_bytes).decode('ascii')
    original_name = secure_filename(file.filename)
    filename = secure_filename(f"resume_{session['user_id']}_{file.filename}")
    db = get_db()
    db.execute(
        'UPDATE candidate_profiles SET resume_filename=?, resume_data=?, resume_original_name=? WHERE user_id=?',
        [filename, encoded, original_name, session['user_id']])
    db.commit()
    if is_ajax:
        return jsonify({'ok': True, 'filename': filename, 'original_name': original_name})
    flash('Resume uploaded successfully!', 'success')
    return redirect(url_for('candidate_profile'))


@app.route('/candidate/profile/delete-resume', methods=['POST'])
@candidate_required
def delete_resume_ajax():
    db = get_db()
    old = db.execute('SELECT resume_filename FROM candidate_profiles WHERE user_id=?',
                     [session['user_id']]).fetchone()
    if old and old.get('resume_filename'):
        try:
            os.remove(os.path.join(RESUME_UPLOAD_FOLDER, old['resume_filename']))
        except OSError:
            pass
    db.execute(
        'UPDATE candidate_profiles SET resume_filename=NULL, resume_data=NULL, resume_original_name=NULL WHERE user_id=?',
        [session['user_id']])
    db.commit()
    return jsonify({'ok': True})


@app.route('/candidate/resume/download')
@candidate_required
def download_resume():
    db = get_db()
    profile = db.execute(
        'SELECT resume_data, resume_original_name, resume_filename FROM candidate_profiles WHERE user_id=?',
        [session['user_id']]
    ).fetchone()
    if not profile or not profile.get('resume_data'):
        flash('Resume not found. Please upload your resume again.', 'error')
        return redirect(url_for('candidate_profile'))
    raw = profile['resume_data']
    if isinstance(raw, memoryview):
        raw = raw.tobytes().decode('ascii')
    file_bytes = base64.b64decode(raw)
    name = profile.get('resume_original_name') or profile.get('resume_filename') or 'resume.pdf'
    ext  = name.rsplit('.', 1)[-1].lower() if '.' in name else 'pdf'
    mime = {
        'pdf':  'application/pdf',
        'doc':  'application/msword',
        'docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
    }.get(ext, 'application/octet-stream')
    as_attachment = request.args.get('dl') == '1'
    return send_file(io.BytesIO(file_bytes), mimetype=mime,
                     as_attachment=as_attachment, download_name=name)


# ── Resume import wizard ──────────────────────────────────────────────────────

def _profile_completion_pct(user_id, db) -> int:
    """Return 0-100 profile completion score for the given candidate."""
    profile = db.execute(
        'SELECT * FROM candidate_profiles WHERE user_id=?', [user_id]
    ).fetchone()
    if not profile:
        return 0
    score = 0
    if profile['headline']  and profile['headline'].strip():   score += 10
    if profile['bio']       and profile['bio'].strip():        score += 10
    if profile['location']  and profile['location'].strip():   score += 8
    if profile['job_title'] and profile['job_title'].strip():  score += 7
    if profile['resume_filename']:                             score += 10
    exp_ct  = db.execute('SELECT COUNT(*) FROM candidate_work_experience WHERE user_id=?', [user_id]).fetchone()[0]
    edu_ct  = db.execute('SELECT COUNT(*) FROM candidate_education       WHERE user_id=?', [user_id]).fetchone()[0]
    sk_ct   = db.execute('SELECT COUNT(*) FROM user_skills               WHERE user_id=?', [user_id]).fetchone()[0]
    proj_ct = db.execute('SELECT COUNT(*) FROM candidate_projects        WHERE user_id=?', [user_id]).fetchone()[0]
    cert_ct = db.execute('SELECT COUNT(*) FROM candidate_certifications  WHERE user_id=?', [user_id]).fetchone()[0]
    if exp_ct  > 0: score += 15
    if edu_ct  > 0: score += 15
    if sk_ct   > 0: score += 10
    if proj_ct > 0: score +=  8
    if cert_ct > 0: score +=  7
    return min(score, 100)


@app.route('/candidate/resume/import')
@candidate_required
def resume_import_page():
    db         = get_db()
    profile    = db.execute('SELECT * FROM candidate_profiles WHERE user_id=?',
                            [session['user_id']]).fetchone()
    completion = _profile_completion_pct(session['user_id'], db)
    return render_template('resume_import.html',
                           user=get_current_user(), profile=profile,
                           completion=completion)


@app.route('/candidate/resume/parse', methods=['POST'])
@candidate_required
def resume_parse_api():
    file = request.files.get('resume')
    if not file or not file.filename:
        return jsonify({'error': 'No file provided.'}), 400

    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in ALLOWED_RESUME_EXT:
        return jsonify({'error': 'Only PDF, DOC, and DOCX resumes up to 1 MB are allowed.'}), 400

    file.seek(0, 2)
    size = file.tell()
    file.seek(0)
    if size > MAX_RESUME_SIZE:
        return jsonify({'error': 'Only PDF, DOC, and DOCX resumes up to 1 MB are allowed.'}), 400

    if not validate_file_magic(file, ALLOWED_RESUME_EXT):
        return jsonify({'error': 'File content does not match the declared format.'}), 400
    file.seek(0)

    os.makedirs(RESUME_UPLOAD_FOLDER, exist_ok=True)
    filename = secure_filename(f"resume_{session['user_id']}_{file.filename}")
    filepath = os.path.join(RESUME_UPLOAD_FOLDER, filename)

    db  = get_db()
    old = db.execute('SELECT resume_filename FROM candidate_profiles WHERE user_id=?',
                     [session['user_id']]).fetchone()

    if old and old['resume_filename'] == filename:
        return jsonify({'error': 'This resume file is already uploaded.'}), 400

    if old and old['resume_filename']:
        try:
            os.remove(os.path.join(RESUME_UPLOAD_FOLDER, old['resume_filename']))
        except OSError:
            pass

    file.save(filepath)
    db.execute('UPDATE candidate_profiles SET resume_filename=? WHERE user_id=?',
               [filename, session['user_id']])
    db.commit()

    # Parse resume text
    try:
        from resume_parser import extract_text, parse_resume as _parse
        text = extract_text(filepath, ext)
        if not text.strip():
            return jsonify({'error': 'Could not read text from this resume. '
                                     'Please try a PDF or DOCX file.'}), 422
        extracted = _parse(text)
    except Exception:
        return jsonify({'error': 'Resume parsing failed. Please try a different file.'}), 422

    # Fetch existing profile data for the comparison UI
    profile    = db.execute('SELECT * FROM candidate_profiles WHERE user_id=?',
                            [session['user_id']]).fetchone()
    experiences = db.execute(
        'SELECT * FROM candidate_work_experience WHERE user_id=?',
        [session['user_id']]).fetchall()
    educations  = db.execute(
        'SELECT * FROM candidate_education WHERE user_id=?',
        [session['user_id']]).fetchall()
    certs       = db.execute(
        'SELECT * FROM candidate_certifications WHERE user_id=?',
        [session['user_id']]).fetchall()
    projects    = db.execute(
        'SELECT * FROM candidate_projects WHERE user_id=?',
        [session['user_id']]).fetchall()
    my_skills   = db.execute(
        'SELECT s.name FROM user_skills us JOIN skills s ON us.skill_id=s.id WHERE us.user_id=?',
        [session['user_id']]).fetchall()

    existing = {
        'headline':        profile['headline']  or '' if profile else '',
        'job_title':       profile['job_title'] or '' if profile else '',
        'experience':      profile['experience'] or '' if profile else '',
        'location':        profile['location']  or '' if profile else '',
        'bio':             profile['bio']        or '' if profile else '',
        'work_experience': [dict(r) for r in experiences],
        'education':       [dict(r) for r in educations],
        'skills':          [r['name'] for r in my_skills],
        'projects':        [dict(r) for r in projects],
        'certifications':  [dict(r) for r in certs],
    }

    return jsonify({'filename': filename, 'extracted': extracted, 'existing': existing})


@app.route('/candidate/resume/confirm', methods=['POST'])
@candidate_required
def resume_confirm():
    data = request.get_json(force=True)
    if not data:
        return jsonify({'error': 'No data provided.'}), 400

    db      = get_db()
    user_id = session['user_id']

    # Profile scalar fields
    allowed_fields = ['headline', 'job_title', 'experience', 'location', 'bio']
    updates = {f: str(data['fields'][f]).strip()
               for f in allowed_fields
               if f in data.get('fields', {}) and data['fields'][f] is not None}
    if updates:
        set_clause = ', '.join(f"{k}=?" for k in updates)
        db.execute(f"UPDATE candidate_profiles SET {set_clause} WHERE user_id=?",
                   list(updates.values()) + [user_id])

    # Work experience (add new entries only)
    for exp in data.get('work_experience', []):
        db.execute(
            '''INSERT INTO candidate_work_experience
               (user_id, company, designation, start_date, end_date, is_current, description)
               VALUES (?, ?, ?, ?, ?, ?, ?)''',
            [user_id,
             str(exp.get('company', ''))[:200],
             str(exp.get('designation', ''))[:200],
             str(exp.get('start_date', '')),
             str(exp.get('end_date', '')),
             1 if exp.get('is_current') else 0,
             str(exp.get('description', ''))[:1000]])

    # Education
    for edu in data.get('education', []):
        db.execute(
            '''INSERT INTO candidate_education (user_id, degree, college, start_year, end_year)
               VALUES (?, ?, ?, ?, ?)''',
            [user_id,
             str(edu.get('degree', ''))[:200],
             str(edu.get('college', ''))[:200],
             str(edu.get('start_year', '')),
             str(edu.get('end_year', ''))])

    # Skills — match against the skills catalogue; silently skip unknowns
    if data.get('skills'):
        all_skills = db.execute('SELECT id, name FROM skills').fetchall()
        skill_map  = {s['name'].lower(): s['id'] for s in all_skills}
        for sk in data['skills']:
            sid = skill_map.get(sk.strip().lower())
            if sid:
                db.execute(
                    'INSERT OR IGNORE INTO user_skills (user_id, skill_id) VALUES (?, ?)',
                    [user_id, sid])

    # Projects
    for proj in data.get('projects', []):
        db.execute(
            '''INSERT INTO candidate_projects
               (user_id, project_name, domain, description, project_url, year)
               VALUES (?, ?, ?, ?, ?, ?)''',
            [user_id,
             str(proj.get('project_name', ''))[:200],
             str(proj.get('domain', ''))[:100],
             str(proj.get('description', ''))[:1000],
             str(proj.get('project_url', ''))[:300],
             str(proj.get('year', ''))])

    # Certifications
    for cert in data.get('certifications', []):
        db.execute(
            '''INSERT INTO candidate_certifications
               (user_id, cert_name, issued_by, year, credential_url)
               VALUES (?, ?, ?, ?, ?)''',
            [user_id,
             str(cert.get('cert_name', ''))[:200],
             str(cert.get('issued_by', ''))[:200],
             str(cert.get('year', '')),
             str(cert.get('credential_url', ''))[:300]])

    db.commit()
    return jsonify({'ok': True, 'completion': _profile_completion_pct(user_id, db)})


@app.route('/candidate/profile/upload-photo', methods=['POST'])
@candidate_required
def candidate_upload_photo():
    db = get_db()
    file = request.files.get('photo')
    if not file or not file.filename:
        return ('No file', 400)
    ext = file.filename.rsplit('.', 1)[-1].lower()
    if ext not in ALLOWED_PHOTO_EXT:
        return ('Only JPG, JPEG, PNG, and WEBP images up to 100 KB are allowed.', 400)
    if len(file.read()) > MAX_PHOTO_SIZE:
        return ('Only JPG, JPEG, PNG, and WEBP images up to 100 KB are allowed.', 400)
    file.seek(0)
    if not validate_file_magic(file, ALLOWED_PHOTO_EXT):
        return ('Only JPG, JPEG, PNG, and WEBP images up to 100 KB are allowed.', 400)
    filename = f"candidate_{session['user_id']}.{ext}"
    mime = _PHOTO_MIME.get(ext, 'image/jpeg')
    file_bytes = file.read()
    encoded = base64.b64encode(file_bytes).decode('ascii')
    db.execute(
        'UPDATE candidate_profiles SET profile_photo=?, profile_photo_data=?, profile_photo_mime=? WHERE user_id=?',
        [filename, encoded, mime, session['user_id']])
    db.commit()
    session['profile_photo'] = filename
    return ('OK', 200)


@app.route('/recruiter/profile/upload-photo', methods=['POST'])
@recruiter_required
def recruiter_upload_photo():
    db = get_db()
    file = request.files.get('photo')
    if not file or not file.filename:
        return ('No file', 400)
    ext = file.filename.rsplit('.', 1)[-1].lower()
    if ext not in ALLOWED_PHOTO_EXT:
        return ('Only JPG, JPEG, PNG, and WEBP images up to 100 KB are allowed.', 400)
    if len(file.read()) > MAX_PHOTO_SIZE:
        return ('Only JPG, JPEG, PNG, and WEBP images up to 100 KB are allowed.', 400)
    file.seek(0)
    if not validate_file_magic(file, ALLOWED_PHOTO_EXT):
        return ('Only JPG, JPEG, PNG, and WEBP images up to 100 KB are allowed.', 400)
    filename = f"recruiter_{session['user_id']}.{ext}"
    mime = _PHOTO_MIME.get(ext, 'image/jpeg')
    file_bytes = file.read()
    encoded = base64.b64encode(file_bytes).decode('ascii')
    db.execute(
        'UPDATE recruiter_profiles SET profile_photo=?, profile_photo_data=?, profile_photo_mime=? WHERE user_id=?',
        [filename, encoded, mime, session['user_id']])
    db.commit()
    session['profile_photo'] = filename
    return ('OK', 200)


@app.route('/candidate/profile/remove-photo', methods=['POST'])
@candidate_required
def candidate_remove_photo():
    db = get_db()
    row = db.execute('SELECT profile_photo FROM candidate_profiles WHERE user_id=?',
                     [session['user_id']]).fetchone()
    db.execute(
        'UPDATE candidate_profiles SET profile_photo=NULL, profile_photo_data=NULL, profile_photo_mime=NULL WHERE user_id=?',
        [session['user_id']])
    db.commit()
    session.pop('profile_photo', None)
    return ('OK', 200)


@app.route('/recruiter/profile/remove-photo', methods=['POST'])
@recruiter_required
def recruiter_remove_photo():
    db = get_db()
    row = db.execute('SELECT profile_photo FROM recruiter_profiles WHERE user_id=?',
                     [session['user_id']]).fetchone()
    db.execute(
        'UPDATE recruiter_profiles SET profile_photo=NULL, profile_photo_data=NULL, profile_photo_mime=NULL WHERE user_id=?',
        [session['user_id']])
    db.commit()
    session.pop('profile_photo', None)
    return ('OK', 200)


@app.route('/candidate/profile/experience/add', methods=['POST'])
@candidate_required
def add_experience():
    company     = request.form.get('company', '').strip()
    designation = request.form.get('designation', '').strip()
    start_date  = request.form.get('start_date', '').strip()
    end_date    = request.form.get('end_date', '').strip()
    is_current  = 1 if request.form.get('is_current') else 0
    description = request.form.get('description', '').strip()
    if not company or not designation:
        flash('Company and designation are required.', 'error')
        return redirect(url_for('candidate_profile'))
    db = get_db()
    if is_current:
        db.execute('UPDATE candidate_work_experience SET is_current=0 WHERE user_id=?',
                   [session['user_id']])
    db.execute('''INSERT INTO candidate_work_experience
                  (user_id, company, designation, start_date, end_date, is_current, description)
                  VALUES (?,?,?,?,?,?,?)''',
               [session['user_id'], company, designation, start_date, end_date, is_current, description])
    db.commit()
    flash('Work experience added.', 'success')
    return redirect(url_for('candidate_profile') + '#section-work-experience')


@app.route('/candidate/profile/experience/<int:exp_id>/edit', methods=['POST'])
@candidate_required
def edit_experience(exp_id):
    company     = request.form.get('company', '').strip()
    designation = request.form.get('designation', '').strip()
    start_date  = request.form.get('start_date', '').strip()
    end_date    = request.form.get('end_date', '').strip()
    is_current  = 1 if request.form.get('is_current') else 0
    description = request.form.get('description', '').strip()
    if not company or not designation:
        flash('Company and designation are required.', 'error')
        return redirect(url_for('candidate_profile') + '#section-work-experience')
    if is_current:
        end_date = ''
    db = get_db()
    if is_current:
        db.execute('UPDATE candidate_work_experience SET is_current=0 WHERE user_id=? AND id!=?',
                   [session['user_id'], exp_id])
    db.execute('''UPDATE candidate_work_experience
                  SET company=?, designation=?, start_date=?, end_date=?, is_current=?, description=?
                  WHERE id=? AND user_id=?''',
               [company, designation, start_date, end_date, is_current, description,
                exp_id, session['user_id']])
    db.commit()
    flash('Work experience updated.', 'success')
    return redirect(url_for('candidate_profile') + '#section-work-experience')


@app.route('/candidate/profile/experience/<int:exp_id>/delete', methods=['POST'])
@candidate_required
def delete_experience(exp_id):
    db = get_db()
    db.execute('DELETE FROM candidate_work_experience WHERE id=? AND user_id=?',
               [exp_id, session['user_id']])
    db.commit()
    flash('Experience removed.', 'success')
    return redirect(url_for('candidate_profile') + '#section-work-experience')


@app.route('/candidate/profile/education/add', methods=['POST'])
@candidate_required
def add_education():
    degree     = request.form.get('degree', '').strip()
    college    = request.form.get('college', '').strip()
    start_year = request.form.get('start_year', '').strip()
    end_year   = request.form.get('end_year', '').strip()
    if not degree or not college:
        flash('Degree and college are required.', 'error')
        return redirect(url_for('candidate_profile') + '#section-education')
    db = get_db()
    db.execute('''INSERT INTO candidate_education (user_id, degree, college, start_year, end_year)
                  VALUES (?,?,?,?,?)''',
               [session['user_id'], degree, college, start_year, end_year])
    db.commit()
    flash('Education added.', 'success')
    return redirect(url_for('candidate_profile') + '#section-education')


@app.route('/candidate/profile/education/<int:edu_id>/edit', methods=['POST'])
@candidate_required
def edit_education(edu_id):
    degree     = request.form.get('degree', '').strip()
    college    = request.form.get('college', '').strip()
    start_year = request.form.get('start_year', '').strip()
    end_year   = request.form.get('end_year', '').strip()
    if not degree or not college:
        flash('Degree and college are required.', 'error')
        return redirect(url_for('candidate_profile') + '#section-education')
    db = get_db()
    db.execute('''UPDATE candidate_education
                  SET degree=?, college=?, start_year=?, end_year=?
                  WHERE id=? AND user_id=?''',
               [degree, college, start_year, end_year, edu_id, session['user_id']])
    db.commit()
    flash('Education updated.', 'success')
    return redirect(url_for('candidate_profile') + '#section-education')


@app.route('/candidate/profile/education/<int:edu_id>/delete', methods=['POST'])
@candidate_required
def delete_education(edu_id):
    db = get_db()
    db.execute('DELETE FROM candidate_education WHERE id=? AND user_id=?',
               [edu_id, session['user_id']])
    db.commit()
    flash('Education removed.', 'success')
    return redirect(url_for('candidate_profile') + '#section-education')


@app.route('/candidate/profile/certification/add', methods=['POST'])
@candidate_required
def add_certification():
    cert_name      = request.form.get('cert_name', '').strip()
    issued_by      = request.form.get('issued_by', '').strip()
    year           = request.form.get('year', '').strip()
    credential_url = request.form.get('credential_url', '').strip()
    if not cert_name:
        flash('Certificate name is required.', 'error')
        return redirect(url_for('candidate_profile') + '#section-certifications')
    db = get_db()
    db.execute('''INSERT INTO candidate_certifications (user_id, cert_name, issued_by, year, credential_url)
                  VALUES (?,?,?,?,?)''',
               [session['user_id'], cert_name, issued_by, year, credential_url])
    db.commit()
    flash('Certification added.', 'success')
    return redirect(url_for('candidate_profile') + '#section-certifications')


@app.route('/candidate/profile/certification/<int:cert_id>/edit', methods=['POST'])
@candidate_required
def edit_certification(cert_id):
    cert_name      = request.form.get('cert_name', '').strip()
    issued_by      = request.form.get('issued_by', '').strip()
    year           = request.form.get('year', '').strip()
    credential_url = request.form.get('credential_url', '').strip()
    if not cert_name:
        flash('Certificate name is required.', 'error')
        return redirect(url_for('candidate_profile') + '#section-certifications')
    db = get_db()
    db.execute('''UPDATE candidate_certifications
                  SET cert_name=?, issued_by=?, year=?, credential_url=?
                  WHERE id=? AND user_id=?''',
               [cert_name, issued_by, year, credential_url, cert_id, session['user_id']])
    db.commit()
    flash('Certification updated.', 'success')
    return redirect(url_for('candidate_profile') + '#section-certifications')


@app.route('/candidate/profile/certification/<int:cert_id>/delete', methods=['POST'])
@candidate_required
def delete_certification(cert_id):
    db = get_db()
    db.execute('DELETE FROM candidate_certifications WHERE id=? AND user_id=?',
               [cert_id, session['user_id']])
    db.commit()
    flash('Certification removed.', 'success')
    return redirect(url_for('candidate_profile') + '#section-certifications')


@app.route('/candidate/profile/project/add', methods=['POST'])
@candidate_required
def add_project():
    project_name = request.form.get('project_name', '').strip()
    domain       = request.form.get('domain', '').strip()
    description  = request.form.get('description', '').strip()
    project_url  = request.form.get('project_url', '').strip()
    year         = request.form.get('year', '').strip()
    if not project_name:
        flash('Project name is required.', 'error')
        return redirect(url_for('candidate_profile') + '#section-projects')
    db = get_db()
    db.execute('''INSERT INTO candidate_projects (user_id, project_name, domain, description, project_url, year)
                  VALUES (?,?,?,?,?,?)''',
               [session['user_id'], project_name, domain, description, project_url, year])
    db.commit()
    flash('Project added.', 'success')
    return redirect(url_for('candidate_profile') + '#section-projects')


@app.route('/candidate/profile/project/<int:proj_id>/edit', methods=['POST'])
@candidate_required
def edit_project(proj_id):
    project_name = request.form.get('project_name', '').strip()
    domain       = request.form.get('domain', '').strip()
    description  = request.form.get('description', '').strip()
    project_url  = request.form.get('project_url', '').strip()
    year         = request.form.get('year', '').strip()
    if not project_name:
        flash('Project name is required.', 'error')
        return redirect(url_for('candidate_profile') + '#section-projects')
    db = get_db()
    db.execute('''UPDATE candidate_projects
                  SET project_name=?, domain=?, description=?, project_url=?, year=?
                  WHERE id=? AND user_id=?''',
               [project_name, domain, description, project_url, year, proj_id, session['user_id']])
    db.commit()
    flash('Project updated.', 'success')
    return redirect(url_for('candidate_profile') + '#section-projects')


@app.route('/candidate/profile/project/<int:proj_id>/delete', methods=['POST'])
@candidate_required
def delete_project(proj_id):
    db = get_db()
    db.execute('DELETE FROM candidate_projects WHERE id=? AND user_id=?',
               [proj_id, session['user_id']])
    db.commit()
    flash('Project removed.', 'success')
    return redirect(url_for('candidate_profile') + '#section-projects')


@app.route('/candidate/profile/phone/send-otp', methods=['POST'])
@candidate_required
def phone_send_otp():
    import random
    code = str(random.randint(100000, 999999))
    session['phone_otp'] = code
    # Demo mode: return code in response (replace with real SMS in production)
    return jsonify({'ok': True, 'code': code, 'demo': True})


@app.route('/candidate/profile/phone/verify-otp', methods=['POST'])
@candidate_required
def phone_verify_otp():
    data = request.get_json(silent=True) or {}
    if data.get('otp') == session.get('phone_otp'):
        session.pop('phone_otp', None)
        phone = data.get('phone', '').strip()
        if phone:
            db = get_db()
            db.execute('UPDATE candidate_profiles SET phone=? WHERE user_id=?',
                       [phone, session['user_id']])
            db.commit()
        return jsonify({'ok': True})
    return jsonify({'ok': False})


@app.route('/candidate/profile/preferences', methods=['POST'])
@candidate_required
def update_preferences():
    work_mode          = request.form.get('work_mode', '').strip()
    notice_period      = request.form.get('notice_period', '').strip()
    expected_salary    = request.form.get('expected_salary', '').strip()
    willing_to_relocate = 1 if request.form.get('willing_to_relocate') else 0
    job_title          = request.form.get('job_title', '').strip()
    location           = request.form.get('location', '').strip()
    db = get_db()
    db.execute('''UPDATE candidate_profiles
                  SET work_mode=?, notice_period=?, expected_salary=?, willing_to_relocate=?,
                      job_title=?, location=?
                  WHERE user_id=?''',
               [work_mode, notice_period, expected_salary, willing_to_relocate,
                job_title, location, session['user_id']])
    db.commit()
    flash('Job preferences updated.', 'success')
    return redirect(url_for('candidate_profile') + '#prefs')


@app.route('/candidate/skills/remove/<int:skill_id>', methods=['POST'])
@candidate_required
def remove_skill(skill_id):
    db = get_db()
    db.execute('DELETE FROM user_skills WHERE user_id=? AND skill_id=?',
               [session['user_id'], skill_id])
    db.commit()
    flash('Skill removed.', 'success')
    return redirect(url_for('candidate_profile'))


# ── Apply ─────────────────────────────────────────────────────────────────────

@app.route('/jobs/<int:job_id>/apply', methods=['POST'])
@candidate_required
def apply_job(job_id):
    db = get_db()
    job = db.execute('''
        SELECT j.*, rp.company AS company_name
        FROM jobs j JOIN recruiter_profiles rp ON j.recruiter_id=rp.user_id
        WHERE j.id=? AND j.active=1
    ''', [job_id]).fetchone()

    if not job:
        return jsonify({'ok': False, 'error': 'job_not_found'}), 404

    if db.execute('SELECT id FROM applications WHERE job_id=? AND candidate_id=?',
                  [job_id, session['user_id']]).fetchone():
        return jsonify({'ok': False, 'error': 'already_applied'}), 409

    cover_letter = request.form.get('cover_letter', '').strip()
    db.execute('INSERT INTO applications (job_id, candidate_id, cover_letter) VALUES (?,?,?)',
               [job_id, session['user_id'], cover_letter])
    db.commit()
    return jsonify({'ok': True})


# ── Recruiter auth ────────────────────────────────────────────────────────────

@app.route('/recruiter/signup', methods=['GET', 'POST'])
@limiter.limit('10 per minute')
def recruiter_signup():
    if session.get('role') == 'recruiter':
        return redirect(url_for('recruiter_dashboard'))

    if request.method == 'POST':
        name             = request.form.get('name', '').strip()
        email            = request.form.get('email', '').strip().lower()
        _phone_code      = request.form.get('phone_code', '').strip()
        _phone_num       = request.form.get('phone', '').strip()
        phone            = ((_phone_code + ' ' + _phone_num).strip()) if _phone_num else ''
        job_title        = request.form.get('job_title', '').strip()
        company          = request.form.get('company', '').strip()
        company_size     = request.form.get('company_size', '').strip()
        industry         = request.form.get('industry', '').strip()
        company_location = request.form.get('company_location', '').strip()
        company_bio      = request.form.get('company_bio', '').strip()
        website          = request.form.get('website', '').strip()
        password         = request.form.get('password', '')
        confirm          = request.form.get('confirm_password', '')
        terms            = request.form.get('terms')

        errors = []
        if not name:    errors.append('Full name is required.')
        if not email:   errors.append('Work email is required.')
        if not company: errors.append('Company name is required.')
        if not terms:   errors.append('You must accept the terms.')

        pw_err = validate_password(password)
        if pw_err:     errors.append(pw_err)
        if not pw_err and password != confirm:
            errors.append('Passwords do not match.')
        if not errors and get_db().execute('SELECT id FROM users WHERE email=?', [email]).fetchone():
            errors.append('An account with this email already exists.')

        if errors:
            for e in errors:
                flash(e, 'error')
            return render_template('recruiter_signup.html', user=None, form=request.form)

        db = get_db()
        raw_token, token_hash, expires_at = _gen_email_token()
        pw_hash = generate_password_hash(password)
        try:
            cur = db.execute(
                'INSERT INTO users (name, email, password_hash, role, email_verified,'
                ' email_verification_token_hash, email_verification_expires_at, email_verification_sent_at)'
                ' VALUES (?,?,?,?,0,?,?,?)',
                [name, email, pw_hash, 'recruiter', token_hash, expires_at, datetime.utcnow()]
            )
        except Exception:
            raw_token = None
            cur = db.execute(
                'INSERT INTO users (name, email, password_hash, role) VALUES (?,?,?,?)',
                [name, email, pw_hash, 'recruiter']
            )
        uid = cur.lastrowid
        db.execute('''INSERT INTO recruiter_profiles
                      (user_id, company, company_bio, website, phone, job_title, company_size, industry, company_location)
                      VALUES (?,?,?,?,?,?,?,?,?)''',
                   [uid, company, company_bio, website, phone, job_title, company_size, industry, company_location])
        db.commit()

        if raw_token:
            send_verification_email(email, name, raw_token)
        session.update({'user_id': uid, 'role': 'recruiter', 'name': name,
                        'email_verified': not raw_token, 'pending_email': email, 'pending_name': name})
        return redirect(url_for('email_sent') if raw_token else url_for('recruiter_dashboard'))

    return render_template('recruiter_signup.html', user=None, form={})


@app.route('/recruiter/login', methods=['GET', 'POST'])
@limiter.limit('15 per minute')
def recruiter_login():
    if session.get('role') == 'recruiter':
        return redirect(url_for('recruiter_dashboard'))

    email = ''
    mismatch = None
    if request.method == 'POST':
        is_ajax = request.is_json
        if is_ajax:
            _d    = request.get_json(silent=True) or {}
            email = _d.get('email', '').strip().lower()
            password = _d.get('password', '')
        else:
            email    = request.form.get('email', '').strip().lower()
            password = request.form.get('password', '')

        _any = get_db().execute("SELECT role FROM users WHERE email=?", [email]).fetchone()
        actual_role = _any['role'] if _any else None
        app.logger.info('[LOGIN:recruiter] email=%r login_type=recruiter actual_role=%r', email, actual_role)

        if _any and actual_role != 'recruiter':
            app.logger.info('[LOGIN:recruiter] role_mismatch=True actual_role=%r', actual_role)
            if is_ajax:
                return jsonify({'success': False, 'role_mismatch': True, 'actual_role': actual_role})
            mismatch = actual_role
            return render_template('recruiter_login.html', user=None, email=email, mismatch=mismatch)

        user = get_db().execute(
            "SELECT * FROM users WHERE email=? AND role='recruiter'", [email]
        ).fetchone()
        app.logger.info('[LOGIN:recruiter] user_found=%s', user is not None)
        pw_ok = bool(user) and check_password_hash(user['password_hash'], password)
        app.logger.info('[LOGIN:recruiter] password_check=%s', pw_ok)
        if pw_ok:
            if not user.get('email_verified', 0):
                app.logger.info('[LOGIN:recruiter] email_verified=false — blocking login')
                if is_ajax:
                    return jsonify({'success': False, 'verification_required': True,
                                    'email_verified': False,
                                    'error': 'Please verify your email before logging in.'})
                flash('Please verify your email before logging in. Check your inbox or resend the verification link.', 'error')
                return render_template('recruiter_login.html', user=None, email=email,
                                       mismatch=None, verification_required=True)
            try:
                rp = get_db().execute(
                    'SELECT profile_photo, company FROM recruiter_profiles WHERE user_id=?',
                    [user['id']]).fetchone()
            except Exception:
                rp = get_db().execute(
                    'SELECT * FROM recruiter_profiles WHERE user_id=?',
                    [user['id']]).fetchone()
            session.update({'user_id': user['id'], 'role': 'recruiter', 'name': user['name'],
                            'email': user['email'],
                            'profile_photo': rp['profile_photo'] if rp else None,
                            'headline': (rp['company'] or '') if rp else '',
                            'email_verified': bool(user.get('email_verified', 0))})
            if is_ajax:
                return jsonify({'success': True, 'redirect': url_for('recruiter_dashboard')})
            flash(f'Welcome back, {user["name"]}!', 'success')
            return redirect(url_for('recruiter_dashboard'))

        if is_ajax:
            return jsonify({'success': False, 'error': 'Invalid email or password.'})
        flash('Invalid email or password.', 'error')

    if not email:
        email = request.args.get('email', request.args.get('oauth_email', ''))
    if mismatch is None:
        mismatch = request.args.get('oauth_mismatch')
    return render_template('recruiter_login.html', user=None, email=email, mismatch=mismatch)


# ── Recruiter dashboard ───────────────────────────────────────────────────────

@app.route('/recruiter/dashboard')
@recruiter_required
def recruiter_dashboard():
    db = get_db()
    profile = db.execute('SELECT * FROM recruiter_profiles WHERE user_id=?',
                         [session['user_id']]).fetchone()
    my_jobs = db.execute('''
        SELECT j.*,
               (SELECT COUNT(*) FROM applications WHERE job_id=j.id) AS app_count,
               (SELECT GROUP_CONCAT(s.name, ',')
                FROM job_skills js JOIN skills s ON js.skill_id=s.id
                WHERE js.job_id=j.id) AS skills_csv
        FROM jobs j WHERE j.recruiter_id=? AND j.active != 2 ORDER BY j.created_at DESC
    ''', [session['user_id']]).fetchall()
    my_jobs = [dict(j, skills_list=(j['skills_csv'] or '').split(',') if j['skills_csv'] else []) for j in my_jobs]

    recent_apps = db.execute('''
        SELECT a.*, j.title AS job_title, u.name AS candidate_name, u.email AS candidate_email
        FROM applications a JOIN jobs j ON a.job_id=j.id
        JOIN users u ON a.candidate_id=u.id
        WHERE j.recruiter_id=? ORDER BY a.applied_at DESC LIMIT 10
    ''', [session['user_id']]).fetchall()

    total_apps = sum(j['app_count'] for j in my_jobs)
    new_apps = db.execute('''
        SELECT COUNT(*) FROM applications a JOIN jobs j ON a.job_id=j.id
        WHERE j.recruiter_id=? AND a.status='applied'
    ''', [session['user_id']]).fetchone()[0]

    under_review_count = db.execute('''
        SELECT COUNT(*) FROM applications a JOIN jobs j ON a.job_id=j.id
        WHERE j.recruiter_id=? AND a.status='reviewing'
    ''', [session['user_id']]).fetchone()[0]

    shortlisted_count = db.execute('''
        SELECT COUNT(*) FROM applications a JOIN jobs j ON a.job_id=j.id
        WHERE j.recruiter_id=? AND a.status='shortlisted'
    ''', [session['user_id']]).fetchone()[0]

    under_review_pct = round((under_review_count / total_apps * 100)) if total_apps else 0
    shortlisted_pct  = round((shortlisted_count  / total_apps * 100)) if total_apps else 0

    stats = {
        'total_jobs': len(my_jobs),
        'active_jobs': sum(1 for j in my_jobs if j['active']),
        'total_applications': total_apps,
        'new_applications': new_apps,
    }
    return render_template('recruiter_dashboard.html',
                           user=get_current_user(), profile=profile,
                           my_jobs=my_jobs, recent_apps=recent_apps, stats=stats,
                           total_apps=total_apps,
                           under_review_count=under_review_count,
                           shortlisted_count=shortlisted_count,
                           under_review_pct=under_review_pct,
                           shortlisted_pct=shortlisted_pct)


@app.route('/recruiter/post-job', methods=['GET', 'POST'])
@recruiter_required
def post_job():
    db = get_db()
    all_skills = db.execute('SELECT * FROM skills ORDER BY category, name').fetchall()
    profile = db.execute('SELECT * FROM recruiter_profiles WHERE user_id=?',
                         [session['user_id']]).fetchone()

    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        location = request.form.get('location', '').strip()
        job_type = request.form.get('job_type', 'Full-time')
        description = request.form.get('description', '').strip()
        requirements = request.form.get('requirements', '').strip()
        salary_min = request.form.get('salary_min', 0, type=int)
        salary_max = request.form.get('salary_max', 0, type=int)
        skill_ids = request.form.getlist('skills', type=int)
        custom_skill_names = [n.strip() for n in request.form.getlist('custom_skill_names') if n.strip()]
        skill_test_mandatory = 1 if request.form.get('skill_test_requirement') == 'mandatory' else 0

        if not all([title, location, description]):
            flash('Title, location, and description are required.', 'error')
            return render_template('post_job.html', user=get_current_user(),
                                   all_skills=all_skills, profile=profile)

        if not skill_ids and not custom_skill_names:
            flash('Please add at least one required skill.', 'error')
            return render_template('post_job.html', user=get_current_user(),
                                   all_skills=all_skills, profile=profile)

        # EMAIL VERIFICATION CHECK DISABLED — re-enable when ready
        # if not db.execute('SELECT email_verified FROM users WHERE id=?', [session['user_id']]).fetchone()['email_verified']:
        #     flash('Please verify your work email before posting a live job.', 'error')
        #     return redirect(url_for('recruiter_dashboard'))

        cur = db.execute('''
            INSERT INTO jobs (recruiter_id, title, company, location, job_type,
                              description, requirements, salary_min, salary_max, skill_test_mandatory)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        ''', [session['user_id'], title, profile['company'], location, job_type,
              description, requirements, salary_min, salary_max, skill_test_mandatory])
        job_id = cur.lastrowid

        for sid in skill_ids:
            db.execute('INSERT OR IGNORE INTO job_skills (job_id, skill_id) VALUES (?,?)',
                       [job_id, sid])

        # Resolve custom skill names — insert if new, then link to job
        for name in custom_skill_names:
            existing = db.execute('SELECT id FROM skills WHERE LOWER(name)=LOWER(?)', [name]).fetchone()
            if existing:
                sid = existing['id']
            else:
                cur2 = db.execute('INSERT INTO skills (name, category) VALUES (?,?)', [name, 'Other'])
                sid = cur2.lastrowid
            db.execute('INSERT OR IGNORE INTO job_skills (job_id, skill_id) VALUES (?,?)', [job_id, sid])

        db.commit()
        flash(f'"{title}" posted successfully!', 'success')
        return redirect(url_for('job_detail', job_id=job_id))

    return render_template('post_job.html', user=get_current_user(),
                           all_skills=all_skills, profile=profile)


@app.route('/recruiter/jobs/<int:job_id>/edit', methods=['GET', 'POST'])
@recruiter_required
def edit_job(job_id):
    db = get_db()
    job = db.execute('SELECT * FROM jobs WHERE id=? AND recruiter_id=?',
                     [job_id, session['user_id']]).fetchone()
    if not job:
        flash('Job not found.', 'error')
        return redirect(url_for('recruiter_dashboard'))

    all_skills = db.execute('SELECT * FROM skills ORDER BY category, name').fetchall()
    profile    = db.execute('SELECT * FROM recruiter_profiles WHERE user_id=?',
                            [session['user_id']]).fetchone()
    job_skills = db.execute('''
        SELECT s.id, s.name FROM job_skills js
        JOIN skills s ON js.skill_id = s.id
        WHERE js.job_id = ?
    ''', [job_id]).fetchall()

    if request.method == 'POST':
        title       = request.form.get('title', '').strip()
        location    = request.form.get('location', '').strip()
        job_type    = request.form.get('job_type', 'Full-time')
        description = request.form.get('description', '').strip()
        requirements = request.form.get('requirements', '').strip()
        salary_min  = request.form.get('salary_min', 0, type=int)
        salary_max  = request.form.get('salary_max', 0, type=int)
        skill_ids   = request.form.getlist('skills', type=int)
        custom_skill_names = [n.strip() for n in request.form.getlist('custom_skill_names') if n.strip()]
        skill_test_mandatory = 1 if request.form.get('skill_test_requirement') == 'mandatory' else 0

        if not all([title, location, description]):
            flash('Title, location, and description are required.', 'error')
            return render_template('post_job.html', user=get_current_user(),
                                   all_skills=all_skills, profile=profile,
                                   edit_job=job, job_skills=job_skills)

        db.execute('''
            UPDATE jobs SET title=?, location=?, job_type=?, description=?,
                            requirements=?, salary_min=?, salary_max=?,
                            skill_test_mandatory=?
            WHERE id=?
        ''', [title, location, job_type, description, requirements,
              salary_min, salary_max, skill_test_mandatory, job_id])

        db.execute('DELETE FROM job_skills WHERE job_id=?', [job_id])
        for sid in skill_ids:
            db.execute('INSERT OR IGNORE INTO job_skills (job_id, skill_id) VALUES (?,?)',
                       [job_id, sid])
        for name in custom_skill_names:
            existing = db.execute('SELECT id FROM skills WHERE LOWER(name)=LOWER(?)', [name]).fetchone()
            sid = existing['id'] if existing else db.execute(
                'INSERT INTO skills (name, category) VALUES (?,?)', [name, 'Other']).lastrowid
            db.execute('INSERT OR IGNORE INTO job_skills (job_id, skill_id) VALUES (?,?)', [job_id, sid])

        db.commit()
        flash(f'"{title}" updated successfully!', 'success')
        return redirect(url_for('job_detail', job_id=job_id))

    return render_template('post_job.html', user=get_current_user(),
                           all_skills=all_skills, profile=profile,
                           edit_job=job, job_skills=job_skills)


@app.route('/recruiter/jobs/<int:job_id>/toggle', methods=['POST'])
@recruiter_required
def toggle_job(job_id):
    db = get_db()
    job = db.execute('SELECT * FROM jobs WHERE id=? AND recruiter_id=?',
                     [job_id, session['user_id']]).fetchone()
    if job:
        db.execute('UPDATE jobs SET active=? WHERE id=?', [0 if job['active'] else 1, job_id])
        db.commit()
        flash('Job status updated.', 'success')
    return redirect(url_for('recruiter_dashboard') + f'#job-{job_id}')


@app.route('/recruiter/jobs/<int:job_id>/archive', methods=['POST'])
@recruiter_required
def archive_job(job_id):
    db = get_db()
    job = db.execute('SELECT * FROM jobs WHERE id=? AND recruiter_id=?',
                     [job_id, session['user_id']]).fetchone()
    if job:
        db.execute('UPDATE jobs SET active=2 WHERE id=?', [job_id])
        db.commit()
        flash(f'"{job["title"]}" has been archived.', 'success')
    return redirect(url_for('recruiter_dashboard') + '#my-job-postings')


@app.route('/recruiter/jobs/<int:job_id>/delete', methods=['POST'])
@recruiter_required
def delete_job(job_id):
    db = get_db()
    job = db.execute('SELECT * FROM jobs WHERE id=? AND recruiter_id=?',
                     [job_id, session['user_id']]).fetchone()
    if job:
        db.execute('DELETE FROM job_skills WHERE job_id=?', [job_id])
        db.execute('DELETE FROM applications WHERE job_id=?', [job_id])
        db.execute('DELETE FROM jobs WHERE id=?', [job_id])
        db.commit()
        flash(f'"{job["title"]}" has been permanently deleted.', 'success')
    return redirect(url_for('recruiter_dashboard') + '#my-job-postings')


@app.route('/recruiter/jobs/<int:job_id>/applications')
@recruiter_required
def job_applications(job_id):
    db = get_db()
    job = db.execute('SELECT * FROM jobs WHERE id=? AND recruiter_id=?',
                     [job_id, session['user_id']]).fetchone()
    if not job:
        flash('Job not found.', 'error')
        return redirect(url_for('recruiter_dashboard'))

    raw_apps = db.execute('''
        SELECT a.*, u.name AS candidate_name, u.email AS candidate_email,
               cp.headline, cp.location AS candidate_location, cp.linkedin, cp.github,
               cp.profile_photo,
               (SELECT GROUP_CONCAT(s.name||':'||us.verified, ',')
                FROM user_skills us JOIN skills s ON us.skill_id=s.id
                WHERE us.user_id=a.candidate_id) AS skills_data
        FROM applications a JOIN users u ON a.candidate_id=u.id
        LEFT JOIN candidate_profiles cp ON a.candidate_id=cp.user_id
        WHERE a.job_id=? ORDER BY a.applied_at DESC
    ''', [job_id]).fetchall()

    applications = []
    for a in raw_apps:
        d = dict(a)
        d['skills'] = parse_skills_data(a['skills_data'])[:8]
        applications.append(d)

    status_counts = {'applied': 0, 'reviewing': 0, 'shortlisted': 0, 'rejected': 0}
    for a in applications:
        if a['status'] in status_counts:
            status_counts[a['status']] += 1

    return render_template('job_applications.html', job=job,
                           applications=applications, user=get_current_user(),
                           status_counts=status_counts)


@app.route('/recruiter/applications/<int:app_id>/status', methods=['POST'])
@recruiter_required
def update_app_status(app_id):
    status = request.form.get('status')
    if status not in ('applied', 'reviewing', 'shortlisted', 'rejected'):
        flash('Invalid status.', 'error')
        return redirect(url_for('recruiter_dashboard'))

    db = get_db()

    # Fetch current application state — need old_status and candidate_id for notification logic
    row = db.execute('''
        SELECT a.job_id, a.candidate_id, a.status AS old_status
        FROM applications a JOIN jobs j ON a.job_id=j.id
        WHERE a.id=? AND j.recruiter_id=?
    ''', [app_id, session['user_id']]).fetchone()

    if not row:
        flash('Application not found.', 'error')
        return redirect(url_for('recruiter_dashboard'))

    try:
        db.execute('UPDATE applications SET status=?, updated_at=CURRENT_TIMESTAMP WHERE id=?',
                   [status, app_id])
    except Exception:
        db.execute('UPDATE applications SET status=? WHERE id=?', [status, app_id])

    # Create shortlisted notification only on first-time transition to shortlisted
    if status == 'shortlisted' and row['old_status'] != 'shortlisted':
        already_notified = db.execute(
            'SELECT id FROM notifications WHERE application_id=? AND type=?',
            [app_id, 'shortlisted']
        ).fetchone()
        if not already_notified:
            job_info = db.execute('''
                SELECT j.title, rp.company
                FROM jobs j JOIN recruiter_profiles rp ON j.recruiter_id=rp.user_id
                WHERE j.id=?
            ''', [row['job_id']]).fetchone()
            if job_info:
                db.execute('''
                    INSERT INTO notifications
                        (candidate_id, job_id, application_id, title, message, type, is_read, redirect_url)
                    VALUES (?, ?, ?, ?, ?, ?, 0, ?)
                ''', [
                    row['candidate_id'],
                    row['job_id'],
                    app_id,
                    'You have been shortlisted!',
                    f'Congratulations! You have been shortlisted for {job_info["title"]} at {job_info["company"]}.',
                    'shortlisted',
                    '/candidate/profile#section-applications',
                ])

    db.commit()
    flash('Application status updated successfully.', 'success')
    return redirect(url_for('job_applications', job_id=row['job_id']))


@app.route('/recruiter/candidates')
@recruiter_required
def search_candidates():
    from datetime import datetime as _dt
    db = get_db()
    search          = request.args.get('q', '').strip()
    skill_filters   = [s.strip() for s in request.args.getlist('skill') if s.strip()]
    location_filter = request.args.get('location', '').strip()
    verified_only   = request.args.get('verified', '') == '1'
    notice_filter   = request.args.get('notice', '').strip()
    work_mode_filter= request.args.get('work_mode', '').strip()
    exp_min         = request.args.get('exp_min', '').strip()
    exp_max         = request.args.get('exp_max', '').strip()
    sort_by         = request.args.get('sort', 'best_match')

    query = '''
        SELECT DISTINCT u.id, u.name, u.email, u.created_at,
               cp.headline, cp.location, cp.profile_photo, cp.experience,
               cp.notice_period, cp.work_mode, cp.work_status,
               (SELECT COUNT(*) FROM user_skills WHERE user_id=u.id AND verified=1) AS verified_count,
               (SELECT COUNT(*) FROM user_skills WHERE user_id=u.id) AS total_skills,
               (SELECT GROUP_CONCAT(s.name||':'||us.verified, ',')
                FROM user_skills us JOIN skills s ON us.skill_id=s.id
                WHERE us.user_id=u.id) AS skills_data
        FROM users u LEFT JOIN candidate_profiles cp ON u.id=cp.user_id
        WHERE u.role='candidate'
    '''
    params = []
    if search:
        query += ''' AND (u.name LIKE ? OR cp.headline LIKE ? OR cp.bio LIKE ?
                    OR u.id IN (
                        SELECT DISTINCT us.user_id FROM user_skills us JOIN skills s ON us.skill_id=s.id
                        WHERE s.name LIKE ?))'''
        params += [f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%']
    if location_filter:
        query += ' AND cp.location LIKE ?'
        params.append(f'%{location_filter}%')
    if skill_filters:
        ph = ','.join('?' * len(skill_filters))
        query += f''' AND u.id IN (
            SELECT DISTINCT us.user_id FROM user_skills us JOIN skills s ON us.skill_id=s.id
            WHERE LOWER(s.name) IN ({ph}))'''
        params.extend(s.lower() for s in skill_filters)
    if verified_only:
        query += ' AND (SELECT COUNT(*) FROM user_skills WHERE user_id=u.id AND verified=1) > 0'
    if notice_filter:
        query += ' AND cp.notice_period = ?'
        params.append(notice_filter)
    if work_mode_filter:
        query += ' AND cp.work_mode = ?'
        params.append(work_mode_filter)
    query += ' ORDER BY verified_count DESC, u.created_at DESC'

    raw = db.execute(query, params).fetchall()

    # Fetch work experience for all candidates in one query
    exp_map = {}
    if raw:
        ids = [c['id'] for c in raw]
        placeholders = ','.join('?' * len(ids))
        all_exps = db.execute(
            f'SELECT user_id, start_date, end_date, is_current FROM candidate_work_experience WHERE user_id IN ({placeholders})',
            ids
        ).fetchall()
        now = _dt.now()
        for exp in all_exps:
            uid = exp['user_id']
            if not exp['start_date']:
                continue
            try:
                sy, sm = int(exp['start_date'][:4]), int(exp['start_date'][5:7])
                if exp['is_current'] or not exp['end_date']:
                    ey, em = now.year, now.month
                else:
                    ey, em = int(exp['end_date'][:4]), int(exp['end_date'][5:7])
                exp_map[uid] = exp_map.get(uid, 0) + max(0, (ey - sy) * 12 + (em - sm))
            except Exception:
                pass

    import re as _re
    candidates = []
    for c in raw:
        d = dict(c)
        d['skills']     = parse_skills_data(c['skills_data'])[:6]
        exp_months = exp_map.get(c['id'], 0)
        if exp_months == 0 and c['experience']:
            m = _re.search(r'(\d+(?:\.\d+)?)', c['experience'])
            if m:
                exp_months = int(float(m.group(1)) * 12)
        d['exp_months'] = exp_months
        score = 30
        if d.get('headline'):      score += 5
        if d.get('location'):      score += 5
        if d.get('profile_photo'): score += 5
        if d.get('work_status'):   score += 5
        score += min(d.get('verified_count', 0) * 12, 45)
        d['match_score'] = min(score, 99)
        candidates.append(d)

    # Python-side experience filter
    if exp_min:
        try:
            candidates = [c for c in candidates if c['exp_months'] >= int(exp_min) * 12]
        except ValueError:
            pass
    if exp_max:
        try:
            candidates = [c for c in candidates if c['exp_months'] <= int(exp_max) * 12]
        except ValueError:
            pass

    # Sort
    _notice_order = {'Immediate': 0, '15 Days': 1, '30 Days': 2, '60 Days': 3, '90 Days': 4}
    if sort_by == 'experience':
        candidates.sort(key=lambda c: c['exp_months'], reverse=True)
    elif sort_by == 'notice':
        candidates.sort(key=lambda c: _notice_order.get(c.get('notice_period') or '', 9))
    elif sort_by in ('recent', 'new'):
        candidates.sort(key=lambda c: c['created_at'] or '', reverse=True)
    else:
        candidates.sort(key=lambda c: c['match_score'], reverse=True)

    has_filters = any([search, skill_filters, location_filter, verified_only,
                       notice_filter, work_mode_filter, exp_min, exp_max])
    all_skills = db.execute('SELECT * FROM skills ORDER BY name').fetchall()
    return render_template('candidates.html', candidates=candidates,
                           all_skills=all_skills, search=search,
                           skill_filters=skill_filters, location_filter=location_filter,
                           verified_only=verified_only, notice_filter=notice_filter,
                           work_mode_filter=work_mode_filter, exp_min=exp_min, exp_max=exp_max,
                           sort_by=sort_by, has_filters=has_filters,
                           user=get_current_user())


@app.route('/recruiter/candidates/export')
@recruiter_required
def export_candidates():
    import csv, io
    db = get_db()
    search          = request.args.get('q', '').strip()
    skill_filters   = [s.strip() for s in request.args.getlist('skill') if s.strip()]
    location_filter = request.args.get('location', '').strip()
    verified_only   = request.args.get('verified', '') == '1'
    notice_filter   = request.args.get('notice', '').strip()
    work_mode_filter= request.args.get('work_mode', '').strip()

    query = '''
        SELECT DISTINCT u.name, u.email, cp.headline, cp.location,
               cp.notice_period, cp.work_mode, cp.work_status,
               (SELECT COUNT(*) FROM user_skills WHERE user_id=u.id AND verified=1) AS verified_skills,
               (SELECT COUNT(*) FROM user_skills WHERE user_id=u.id) AS total_skills
        FROM users u LEFT JOIN candidate_profiles cp ON u.id=cp.user_id
        WHERE u.role='candidate'
    '''
    params = []
    if search:
        query += ' AND (u.name LIKE ? OR cp.headline LIKE ?)'
        params += [f'%{search}%', f'%{search}%']
    if location_filter:
        query += ' AND cp.location LIKE ?'
        params.append(f'%{location_filter}%')
    if skill_filters:
        ph = ','.join('?' * len(skill_filters))
        query += f' AND u.id IN (SELECT DISTINCT us.user_id FROM user_skills us JOIN skills s ON us.skill_id=s.id WHERE s.name IN ({ph}))'
        params.extend(skill_filters)
    if verified_only:
        query += ' AND (SELECT COUNT(*) FROM user_skills WHERE user_id=u.id AND verified=1) > 0'
    if notice_filter:
        query += ' AND cp.notice_period = ?'
        params.append(notice_filter)
    if work_mode_filter:
        query += ' AND cp.work_mode = ?'
        params.append(work_mode_filter)
    query += ' ORDER BY u.name'

    rows = db.execute(query, params).fetchall()
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(['Name', 'Email', 'Headline', 'Location', 'Notice Period', 'Availability', 'Status', 'Verified Skills', 'Total Skills'])
    for r in rows:
        w.writerow([r['name'], r['email'], r['headline'] or '', r['location'] or '',
                    r['notice_period'] or '', r['work_mode'] or '', r['work_status'] or '',
                    r['verified_skills'], r['total_skills']])
    out.seek(0)
    return Response(out.getvalue(), mimetype='text/csv',
                    headers={'Content-Disposition': 'attachment;filename=candidates.csv'})


# ── Candidate Detail ─────────────────────────────────────────────────────────

@app.route('/candidates/<int:candidate_id>')
def candidate_detail(candidate_id):
    db = get_db()

    if session.get('role') == 'recruiter':
        viewer = db.execute('SELECT email_verified FROM users WHERE id=?', [session['user_id']]).fetchone()
        if viewer and not viewer.get('email_verified', 0):
            return render_template('candidate_detail.html', candidate=None, skills=[],
                                   blocked_recruiter=True, user=get_current_user())

    candidate = db.execute('''
        SELECT u.id, u.name, u.email, u.created_at, u.email_verified,
               cp.headline, cp.location, cp.bio, cp.linkedin, cp.github,
               cp.phone, cp.gender, cp.differently_abled
        FROM users u LEFT JOIN candidate_profiles cp ON u.id = cp.user_id
        WHERE u.id = ? AND u.role = 'candidate'
    ''', [candidate_id]).fetchone()
    if not candidate:
        return render_template('candidate_detail.html', candidate=None, skills=[],
                               blocked_recruiter=False, user=get_current_user())
    skills = db.execute('''
        SELECT s.id, s.name, s.category, us.verified, us.score
        FROM user_skills us JOIN skills s ON us.skill_id = s.id
        WHERE us.user_id = ?
        ORDER BY us.verified DESC, s.name
    ''', [candidate_id]).fetchall()
    show_verified_skills = bool(candidate['email_verified']) or (session.get('user_id') == candidate_id)

    experiences = db.execute(
        'SELECT * FROM candidate_work_experience WHERE user_id=? ORDER BY start_date DESC, id DESC',
        [candidate_id]).fetchall()
    educations = db.execute(
        'SELECT * FROM candidate_education WHERE user_id=? ORDER BY COALESCE(end_year,\'9999\') DESC, COALESCE(start_year,\'0\') DESC',
        [candidate_id]).fetchall()
    certifications = db.execute(
        'SELECT * FROM candidate_certifications WHERE user_id=? ORDER BY COALESCE(year,\'0\') DESC, id DESC',
        [candidate_id]).fetchall()
    try:
        projects = db.execute(
            'SELECT * FROM candidate_projects WHERE user_id=? ORDER BY COALESCE(year,\'0\') DESC, id DESC',
            [candidate_id]).fetchall()
    except Exception:
        projects = db.execute(
            'SELECT * FROM candidate_projects WHERE user_id=? ORDER BY id DESC',
            [candidate_id]).fetchall()
    profile_extra = db.execute(
        'SELECT resume_filename, profile_photo FROM candidate_profiles WHERE user_id=?',
        [candidate_id]).fetchone()

    now = datetime.now()
    total_months = 0
    for exp in experiences:
        if not exp['start_date']:
            continue
        try:
            sy, sm = int(exp['start_date'][:4]), int(exp['start_date'][5:7])
        except (ValueError, IndexError):
            continue
        if exp['is_current']:
            ey, em = now.year, now.month
        elif exp['end_date']:
            try:
                ey, em = int(exp['end_date'][:4]), int(exp['end_date'][5:7])
            except (ValueError, IndexError):
                continue
        else:
            continue
        diff = (ey - sy) * 12 + (em - sm)
        if diff > 0:
            total_months += diff
    total_exp_years  = total_months // 12
    total_exp_months = total_months % 12

    return render_template('candidate_detail.html', candidate=candidate, skills=skills,
                           blocked_recruiter=False, show_verified_skills=show_verified_skills,
                           experiences=experiences, educations=educations,
                           certifications=certifications, projects=projects,
                           profile_extra=profile_extra,
                           total_exp_years=total_exp_years, total_exp_months=total_exp_months,
                           user=get_current_user())


# ── Recruiter Detail ──────────────────────────────────────────────────────────

@app.route('/recruiters/<int:recruiter_id>')
def recruiter_detail(recruiter_id):
    db = get_db()
    recruiter = db.execute('''
        SELECT u.id, u.name, u.email, u.created_at,
               rp.company, rp.company_bio, rp.website, rp.profile_photo
        FROM users u LEFT JOIN recruiter_profiles rp ON u.id = rp.user_id
        WHERE u.id = ? AND u.role = 'recruiter'
    ''', [recruiter_id]).fetchone()
    if not recruiter:
        return render_template('recruiter_detail.html', recruiter=None, jobs=[],
                               user=get_current_user())
    jobs = db.execute('''
        SELECT j.*,
               (SELECT GROUP_CONCAT(s.name, ', ')
                FROM job_skills js JOIN skills s ON js.skill_id = s.id
                WHERE js.job_id = j.id) AS skills_list
        FROM jobs j WHERE j.recruiter_id = ? AND j.active = 1
        ORDER BY j.created_at DESC
    ''', [recruiter_id]).fetchall()
    return render_template('recruiter_detail.html', recruiter=recruiter, jobs=jobs,
                           user=get_current_user())


# ── Email verification ────────────────────────────────────────────────────────

@app.route('/email-sent')
def email_sent():
    pending_email = session.get('pending_email', '')
    pending_name  = session.get('pending_name', '')
    if not pending_email:
        return redirect(url_for('home'))
    sent_at_ts = 0
    try:
        db  = get_db()
        row = db.execute(
            'SELECT email_verification_sent_at FROM users WHERE email=?', [pending_email]
        ).fetchone()
        if row and row['email_verification_sent_at']:
            sent_dt = row['email_verification_sent_at']
            if isinstance(sent_dt, str):
                sent_dt = datetime.strptime(sent_dt[:19], '%Y-%m-%d %H:%M:%S')
            sent_at_ts = int(sent_dt.timestamp())
    except Exception:
        pass
    return render_template('email_sent.html', user=get_current_user(),
                           pending_email=pending_email, pending_name=pending_name,
                           sent_at_ts=sent_at_ts)


@app.route('/verify-email/<token>')
def verify_email(token):
    token_hash = hashlib.sha256(token.encode()).hexdigest()
    db = get_db()
    try:
        user = db.execute(
            'SELECT * FROM users WHERE email_verification_token_hash=?', [token_hash]
        ).fetchone()
    except Exception:
        flash('Verification is temporarily unavailable. Please try again shortly.', 'error')
        return redirect(url_for('home'))
    if not user:
        flash('This verification link is invalid or has already been used. Please request a new one.', 'error')
        return redirect(url_for('home'))
    if user.get('email_verified', 0):
        flash('Your email is already verified. Please log in.', 'success')
        session.pop('pending_email', None)
        session.pop('pending_name', None)
        session.update({'user_id': user['id'], 'role': user['role'], 'name': user['name'],
                        'email_verified': True})
        return redirect(url_for('recruiter_dashboard') if user['role'] == 'recruiter' else url_for('candidate_profile'))
    expires_at = user.get('email_verification_expires_at')
    if expires_at:
        try:
            if isinstance(expires_at, str):
                expires_at = datetime.strptime(expires_at[:19], '%Y-%m-%d %H:%M:%S')
            if datetime.utcnow() > expires_at:
                flash('This verification link has expired. Please request a new one.', 'error')
                return redirect(url_for('home'))
        except Exception:
            pass
    db.execute(
        'UPDATE users SET email_verified=1, email_verified_at=?, email_verification_token_hash=NULL,'
        ' email_verification_expires_at=NULL WHERE id=?',
        [datetime.utcnow(), user['id']]
    )
    db.commit()
    flash('Email verified! Your account is now active.', 'success')
    session.pop('pending_email', None)
    session.pop('pending_name', None)
    session.update({'user_id': user['id'], 'role': user['role'], 'name': user['name'],
                    'email_verified': True})
    if user['role'] == 'recruiter':
        return redirect(url_for('recruiter_dashboard'))
    return redirect(url_for('candidate_profile'))


@app.route('/resend-verification', methods=['POST'])
@limiter.limit('5 per hour')
def resend_verification():
    db = get_db()

    # Resolve user — either logged in or coming from email_sent page
    if session.get('user_id'):
        user = db.execute('SELECT * FROM users WHERE id=?', [session['user_id']]).fetchone()
    else:
        email = session.get('pending_email', '').strip().lower()
        user = db.execute('SELECT * FROM users WHERE email=?', [email]).fetchone() if email else None

    if not user:
        flash('No pending verification found. Please sign up again.', 'error')
        return redirect(url_for('home'))
    if user.get('email_verified', 0):
        flash('Your email is already verified.', 'success')
        dest = url_for('candidate_profile') if user['role'] == 'candidate' else url_for('recruiter_dashboard')
        return redirect(dest)

    # 60-second cooldown check (column may not exist if migration hasn't run yet)
    try:
        sent_at = user['email_verification_sent_at']
    except (KeyError, IndexError):
        sent_at = None
    if sent_at:
        try:
            if isinstance(sent_at, str):
                sent_at = datetime.strptime(sent_at[:19], '%Y-%m-%d %H:%M:%S')
            seconds_ago = (datetime.utcnow() - sent_at).total_seconds()
            if seconds_ago < 60:
                wait = int(60 - seconds_ago)
                flash(f'Please wait {wait} seconds before requesting another verification email.', 'error')
                session['pending_email'] = user['email']
                session['pending_name'] = user['name']
                return redirect(url_for('email_sent'))
        except Exception:
            pass

    raw_token, token_hash, expires_at = _gen_email_token()
    db.execute(
        'UPDATE users SET email_verification_token_hash=?, email_verification_expires_at=?,'
        ' email_verification_sent_at=? WHERE id=?',
        [token_hash, expires_at, datetime.utcnow(), user['id']]
    )
    db.commit()
    send_verification_email(user['email'], user['name'], raw_token)
    flash('Verification email sent! Please check your inbox.', 'success')
    session['pending_email'] = user['email']
    session['pending_name'] = user['name']
    return redirect(url_for('email_sent'))


# ── Forgot / Reset Password ───────────────────────────────────────────────────

@app.route('/forgot-password', methods=['GET', 'POST'])
@limiter.limit('6 per minute')
def forgot_password():
    role = request.args.get('role', request.form.get('role', 'candidate'))
    if role not in ('candidate', 'recruiter'):
        role = 'candidate'

    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        if email:
            db = get_db()
            user = db.execute(
                'SELECT id, name, email FROM users WHERE email=? AND role=?', [email, role]
            ).fetchone()
            if user:
                raw = secrets.token_urlsafe(32)
                h   = hashlib.sha256(raw.encode()).hexdigest()
                expires = datetime.utcnow() + timedelta(minutes=15)
                reset_route = 'candidate_reset_password' if role == 'candidate' else 'recruiter_reset_password'
                try:
                    db.execute(
                        'UPDATE users SET password_reset_token_hash=?, password_reset_expires_at=? WHERE id=?',
                        [h, expires, user['id']]
                    )
                    db.commit()
                    reset_url = url_for(reset_route, token=raw, _external=True)
                    app.logger.info('[FORGOT_PW] token generated for user_id=%s role=%s', user['id'], role)
                    send_password_reset_email(user['email'], user['name'], reset_url)
                except Exception as exc:
                    app.logger.error('[FORGOT_PW] DB/email error: %s', exc)
            else:
                app.logger.info('[FORGOT_PW] no %s account found for email=%r', role, email)
        # Always show success — never reveal whether email exists
        return render_template('forgot_password.html', sent=True, role=role)

    return render_template('forgot_password.html', sent=False, role=role)


def _handle_reset_password(role):
    """Shared reset-password handler for candidate and recruiter routes."""
    token = request.args.get('token', '') if request.method == 'GET' else request.form.get('token', '')
    login_route = 'candidate_login' if role == 'candidate' else 'recruiter_login'

    if not token:
        return render_template('reset_password.html', invalid=True, role=role, login_route=login_route)

    token_hash = hashlib.sha256(token.encode()).hexdigest()
    db = get_db()
    user = db.execute(
        'SELECT id, name, email, password_reset_token_hash, password_reset_expires_at'
        ' FROM users WHERE password_reset_token_hash=? AND role=?', [token_hash, role]
    ).fetchone()

    now = datetime.utcnow()
    if not user:
        return render_template('reset_password.html', invalid=True, role=role, login_route=login_route)

    expires_raw = user['password_reset_expires_at']
    if isinstance(expires_raw, str):
        try:
            expires_dt = datetime.fromisoformat(expires_raw)
        except ValueError:
            expires_dt = now
    else:
        expires_dt = expires_raw or now

    if expires_dt < now:
        return render_template('reset_password.html', expired=True, role=role, login_route=login_route)

    if request.method == 'POST':
        new_pw  = request.form.get('password', '')
        confirm = request.form.get('confirm_password', '')
        pw_err  = validate_password(new_pw)
        if pw_err:
            return render_template('reset_password.html', token=token, role=role, login_route=login_route, error=pw_err)
        if new_pw != confirm:
            return render_template('reset_password.html', token=token, role=role, login_route=login_route,
                                   error='Passwords do not match.')
        try:
            db.execute(
                'UPDATE users SET password_hash=?, password_reset_token_hash=NULL,'
                ' password_reset_expires_at=NULL WHERE id=?',
                [generate_password_hash(new_pw), user['id']]
            )
            db.commit()
            app.logger.info('[RESET_PW] password updated for user_id=%s role=%s', user['id'], role)
        except Exception as exc:
            app.logger.error('[RESET_PW] DB error: %s', exc)
            return render_template('reset_password.html', token=token, role=role, login_route=login_route,
                                   error='Something went wrong. Please try again.')
        return render_template('reset_password.html', success=True, role=role, login_route=login_route)

    return render_template('reset_password.html', token=token, role=role, login_route=login_route)


@app.route('/candidate/reset-password', methods=['GET', 'POST'])
def candidate_reset_password():
    return _handle_reset_password('candidate')


@app.route('/recruiter/reset-password', methods=['GET', 'POST'])
def recruiter_reset_password():
    return _handle_reset_password('recruiter')


# ── Logout ────────────────────────────────────────────────────────────────────

@app.route('/logout', methods=['POST'])
def logout():
    session.clear()
    flash('You have been logged out.', 'success')
    return redirect(url_for('home'))


# ── Run ───────────────────────────────────────────────────────────────────────

# Auto-initialize SQLite schema when running under gunicorn (not __main__)
if not _USE_POSTGRES:
    try:
        init_db()
    except Exception as _e:
        print(f'[STARTUP] SQLite init error: {_e}')

if __name__ == '__main__':
    app.run(debug=True, host='127.0.0.1', port=5000)
